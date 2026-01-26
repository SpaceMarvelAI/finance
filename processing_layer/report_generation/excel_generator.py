"""
Excel Report Generator
Converts structured data to professional Excel reports
"""

from pathlib import Path
from datetime import datetime
from typing import Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from shared.config.logging_config import get_logger


logger = get_logger(__name__)


class ExcelReportGenerator:
    """
    Excel Report Generator
    Creates professional Excel reports from structured data
    """
    
    def __init__(self, output_dir: str = "./output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Styling
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_ar_aging_excel(self, report_data: Dict[str, Any]) -> str:
        """
        Generate AR Aging Report in Excel
        
        Args:
            report_data: Structured report data
            
        Returns:
            Path to generated Excel file
        """
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AR Aging"
        
        # Title
        ws['A1'] = report_data['report_metadata']['report_name']
        ws['A1'].font = self.title_font
        
        ws['A2'] = f"As of: {report_data['report_metadata']['as_of_date']}"
        ws['A3'] = f"Generated: {report_data['report_metadata']['generated_at']}"
        
        # Summary Section
        row = 5
        ws[f'A{row}'] = "SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        summary = report_data['summary']
        ws[f'A{row}'] = "Total Receivables:"
        ws[f'B{row}'] = f"${summary['total_receivables']:,.2f}"
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Total Overdue:"
        ws[f'B{row}'] = f"${summary['total_overdue']:,.2f}"
        if summary['total_overdue'] > 0:
            ws[f'B{row}'].font = Font(color="FF0000", bold=True)
        row += 1
        
        ws[f'A{row}'] = "Document Count:"
        ws[f'B{row}'] = summary['document_count']
        row += 2
        
        # Aging Buckets Table
        ws[f'A{row}'] = "AGING BREAKDOWN"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Headers
        headers = ["Bucket", "Days", "Amount", "Count", "Percentage"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        row += 1
        
        # Data rows
        for bucket in report_data['aging_buckets']:
            ws.cell(row=row, column=1).value = bucket['bucket']
            ws.cell(row=row, column=2).value = bucket['days']
            ws.cell(row=row, column=3).value = f"${bucket['amount']:,.2f}"
            ws.cell(row=row, column=4).value = bucket['count']
            ws.cell(row=row, column=5).value = f"{bucket['percentage']}%"
            
            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        # Details Section
        if report_data['details']:
            row += 2
            ws[f'A{row}'] = "INVOICE DETAILS"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            # Headers
            detail_headers = ["Document #", "Vendor", "Due Date", "Days Overdue", "Amount", "Bucket"]
            for col, header in enumerate(detail_headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            row += 1
            
            # Data rows
            for detail in report_data['details']:
                ws.cell(row=row, column=1).value = detail['document_number']
                ws.cell(row=row, column=2).value = detail.get('vendor', 'N/A')
                ws.cell(row=row, column=3).value = detail['due_date']
                ws.cell(row=row, column=4).value = detail['days_overdue']
                ws.cell(row=row, column=5).value = f"${detail['amount']:,.2f}"
                ws.cell(row=row, column=6).value = detail['aging_bucket']
                
                # Apply borders
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = self.border
                
                row += 1
        
        # Auto-size columns
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # Save
        filename = f"AR_Aging_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        logger.info(f"AR Aging Excel generated: {filepath}")
        return str(filepath)
    
    def generate_ap_aging_excel(self, report_data: Dict[str, Any]) -> str:
        """Generate AP Aging Report in Excel"""
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AP Aging"
        
        # Title
        ws['A1'] = report_data['report_metadata']['report_name']
        ws['A1'].font = self.title_font
        
        ws['A2'] = f"As of: {report_data['report_metadata']['as_of_date']}"
        ws['A3'] = f"Generated: {report_data['report_metadata']['generated_at']}"
        
        # Summary Section
        row = 5
        ws[f'A{row}'] = "SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        summary = report_data['summary']
        ws[f'A{row}'] = "Total Payables:"
        ws[f'B{row}'] = f"${summary['total_payables']:,.2f}"
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Total Overdue:"
        ws[f'B{row}'] = f"${summary['total_overdue']:,.2f}"
        if summary['total_overdue'] > 0:
            ws[f'B{row}'].font = Font(color="FF0000", bold=True)
        row += 1
        
        ws[f'A{row}'] = "Document Count:"
        ws[f'B{row}'] = summary['document_count']
        row += 2
        
        # Aging Buckets Table
        ws[f'A{row}'] = "AGING BREAKDOWN"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Headers
        headers = ["Bucket", "Days", "Amount", "Count", "Percentage"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        row += 1
        
        # Data rows
        for bucket in report_data['aging_buckets']:
            ws.cell(row=row, column=1).value = bucket['bucket']
            ws.cell(row=row, column=2).value = bucket['days']
            ws.cell(row=row, column=3).value = f"${bucket['amount']:,.2f}"
            ws.cell(row=row, column=4).value = bucket['count']
            ws.cell(row=row, column=5).value = f"{bucket['percentage']}%"
            
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        # Auto-size columns
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # Save
        filename = f"AP_Aging_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        logger.info(f"AP Aging Excel generated: {filepath}")
        return str(filepath)
    
    def generate_vendor_summary_excel(self, report_data: Dict[str, Any]) -> str:
        """Generate Vendor Summary in Excel"""
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vendor Summary"
        
        # Title
        ws['A1'] = report_data['report_metadata']['report_name']
        ws['A1'].font = self.title_font
        
        ws['A2'] = f"Generated: {report_data['report_metadata']['generated_at']}"
        
        # Summary
        row = 4
        summary = report_data['summary']
        ws[f'A{row}'] = "Total Vendors:"
        ws[f'B{row}'] = summary['total_vendors']
        row += 1
        
        ws[f'A{row}'] = "Total Purchase Amount:"
        ws[f'B{row}'] = f"${summary['total_purchase_amount']:,.2f}"
        row += 1
        
        ws[f'A{row}'] = "Total Balance Due:"
        ws[f'B{row}'] = f"${summary['total_balance_due']:,.2f}"
        row += 2
        
        # Vendor Table
        ws[f'A{row}'] = "VENDOR DETAILS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Headers
        headers = ["Vendor Name", "Invoice Count", "Total Amount", "Balance Due", "Paid Amount"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        row += 1
        
        # Data rows
        for vendor in report_data['vendors']:
            ws.cell(row=row, column=1).value = vendor['vendor_name']
            ws.cell(row=row, column=2).value = vendor['invoice_count']
            ws.cell(row=row, column=3).value = f"${vendor['total_amount']:,.2f}"
            ws.cell(row=row, column=4).value = f"${vendor['balance_due']:,.2f}"
            ws.cell(row=row, column=5).value = f"${vendor['paid_amount']:,.2f}"
            
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        # Auto-size columns
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Save
        filename = f"Vendor_Summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        logger.info(f"Vendor Summary Excel generated: {filepath}")
        return str(filepath)