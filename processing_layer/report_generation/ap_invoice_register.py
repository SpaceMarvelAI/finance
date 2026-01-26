"""
AP INVOICE REGISTER - COMPLETE FIXED VERSION
All issues resolved:
- Uses INR amounts from database
- Logo centered at column B
- Company name properly centered
- ₹ symbol (not $)
- Proper descriptions
- Correct invoice numbers
"""

from typing import Dict, Any, Optional, List
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from shared.config.logging_config import get_logger


logger = get_logger(__name__)


# =============================================================================
# REPORT GENERATOR (Data from Database)
# =============================================================================

class APInvoiceRegisterGenerator:
    """Generate AP Invoice Register data from database"""
    
    def __init__(self, mcp_tools=None):
        self.mcp_tools = mcp_tools
        from data_layer.database.database_manager import get_database
        self.db = get_database()
        self.logger = logger
    
    def generate_report(self, company_id: Optional[str] = None) -> Dict[str, Any]:
        """Generate report data"""
        
        # Get purchase invoices from database
        try:
            purchase_docs = self.db.get_documents_by_category('purchase', company_id)
            self.logger.info(f"Retrieved {len(purchase_docs)} purchase documents")
        except Exception as e:
            self.logger.error(f"Database error: {e}")
            purchase_docs = []
        
        invoices = []
        totals = {
            "invoice_amt": 0.0,
            "tax_amt": 0.0,
            "net_amt": 0.0,
            "paid_amt": 0.0,
            "outstanding": 0.0
        }
        
        trans_id = 1001
        
        for doc_entry in purchase_docs:
            # Get INR amounts (already converted)
            invoice_amt = float(doc_entry.get("inr_amount") or 0)
            tax_amt = float(doc_entry.get("tax_total") or 0)
            paid_amt = float(doc_entry.get("paid_amount") or 0)
            
            # Get invoice details
            vendor_name = doc_entry.get("vendor_name") or "Unknown"
            invoice_no = doc_entry.get("document_number") or "N/A"
            invoice_date = doc_entry.get("document_date") or ""
            
            # Get original currency for reference
            original_currency = doc_entry.get("original_currency") or "INR"
            original_amount = float(doc_entry.get("original_amount") or 0)
            
            # Description: Show original currency if not INR
            if original_currency != "INR":
                description = f"{original_currency} {original_amount:.2f}"
            else:
                description = "General Invoice"
            
            # Vendor ID
            vendor_id = self._generate_vendor_id(vendor_name)
            
            # Calculations
            net_amt = invoice_amt - tax_amt
            outstanding = invoice_amt - paid_amt
            
            # Status
            if paid_amt >= invoice_amt:
                status = "Paid"
            elif paid_amt > 0:
                status = "Partial"
            else:
                status = "Unpaid"
            
            # Format dates
            inv_date_str = self._format_date(invoice_date)
            due_date_str = self._format_date(invoice_date)  # Use same date
            
            # Create row
            invoice_row = {
                "trans_id": trans_id,
                "vendor_name": vendor_name,
                "vendor_id": vendor_id,
                "invoice_no": invoice_no,
                "invoice_date": inv_date_str,
                "due_date": due_date_str,
                "description": description,
                "invoice_amt": round(invoice_amt, 2),
                "tax_amt": round(tax_amt, 2),
                "net_amt": round(net_amt, 2),
                "paid_amt": round(paid_amt, 2),
                "outstanding": round(outstanding, 2),
                "status": status
            }
            
            invoices.append(invoice_row)
            
            # Update totals
            totals["invoice_amt"] += invoice_amt
            totals["tax_amt"] += tax_amt
            totals["net_amt"] += net_amt
            totals["paid_amt"] += paid_amt
            totals["outstanding"] += outstanding
            
            trans_id += 1
        
        # Round totals
        for key in totals:
            totals[key] = round(totals[key], 2)
        
        return {
            "report_metadata": {
                "report_type": "AP_INVOICE_REGISTER",
                "generated_at": datetime.now().isoformat(),
                "currency": "INR"
            },
            "summary": {
                "total_invoices": len(invoices),
                "paid_count": sum(1 for inv in invoices if inv["status"] == "Paid"),
                "partial_count": sum(1 for inv in invoices if inv["status"] == "Partial"),
                "unpaid_count": sum(1 for inv in invoices if inv["status"] == "Unpaid")
            },
            "invoices": invoices,
            "totals": totals
        }
    
    def _generate_vendor_id(self, vendor_name: str) -> str:
        """Generate vendor ID"""
        if not vendor_name or vendor_name == "Unknown":
            return "V-000"
        prefix = vendor_name[:3].upper().replace(" ", "")
        number = abs(hash(vendor_name)) % 1000
        return f"V-{prefix}{number:03d}"
    
    def _format_date(self, date_str: str) -> str:
        """Format date to MM/DD/YY"""
        if not date_str:
            return ""
        try:
            dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
            return dt.strftime("%m/%d/%y")
        except:
            return str(date_str)[:10]


# =============================================================================
# EXCEL GENERATOR (Branded Output)
# =============================================================================

class APInvoiceRegisterExcelGenerator:
    """Generate branded Excel file"""
    
    def __init__(self, user_id: str, output_dir: str = "./output/reports"):
        self.user_id = user_id
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Load branding
        from shared.branding.company_branding import CompanyBrandingManager
        self.branding_manager = CompanyBrandingManager()
        self.branding = self.branding_manager.get_branding(user_id)
        
        if not self.branding:
            raise ValueError(f"No branding for user: {user_id}")
        
        colors = self.branding["colors"]
        self.primary_color = colors["primary"].lstrip('#')
        self.secondary_color = colors["secondary"].lstrip('#')
        self.accent_color = colors["accent"].lstrip('#')
        
        self.header_font = Font(bold=True, size=11, color="FFFFFF")
        self.header_fill = PatternFill(
            start_color=self.primary_color,
            end_color=self.primary_color,
            fill_type="solid"
        )
        self.title_font = Font(bold=True, size=16, color=self.primary_color)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.logger = logger
    
    def generate(self, report_data: Dict[str, Any]) -> str:
        """Generate Excel file"""
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AP Invoice Register"
        
        current_row = 1
        
        # Company header (FIXED: Centered logo and company name)
        current_row = self._add_company_header(ws, current_row)
        
        # Title
        current_row += 1
        ws.merge_cells(f'A{current_row}:M{current_row}')
        ws[f'A{current_row}'].value = "AP INVOICE REGISTER REPORT"
        ws[f'A{current_row}'].font = self.title_font
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:M{current_row}')
        ws[f'A{current_row}'].value = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        ws[f'A{current_row}'].font = Font(size=9)
        
        # Summary
        current_row += 2
        summary = report_data['summary']
        ws[f'A{current_row}'] = f"Total Invoices: {summary['total_invoices']}"
        ws[f'D{current_row}'] = f"Paid: {summary['paid_count']}"
        ws[f'G{current_row}'] = f"Partial: {summary['partial_count']}"
        ws[f'J{current_row}'] = f"Unpaid: {summary['unpaid_count']}"
        
        # Headers
        current_row += 2
        headers = [
            "Trans ID", "Vendor Name", "Vendor ID", "Invoice No.",
            "Invoice Date", "Due Date", "Description",
            "Invoice Amt", "Tax Amt", "Net Amt", "Paid Amt",
            "Outstanding", "Status"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        current_row += 1
        
        # Data rows (FIXED: Using ₹ symbol)
        for invoice in report_data['invoices']:
            ws.cell(row=current_row, column=1).value = invoice['trans_id']
            ws.cell(row=current_row, column=2).value = invoice['vendor_name']
            ws.cell(row=current_row, column=3).value = invoice['vendor_id']
            ws.cell(row=current_row, column=4).value = invoice['invoice_no']
            ws.cell(row=current_row, column=5).value = invoice['invoice_date']
            ws.cell(row=current_row, column=6).value = invoice['due_date']
            ws.cell(row=current_row, column=7).value = invoice['description']
            ws.cell(row=current_row, column=8).value = f"₹{invoice['invoice_amt']:,.2f}"
            ws.cell(row=current_row, column=9).value = f"₹{invoice['tax_amt']:,.2f}"
            ws.cell(row=current_row, column=10).value = f"₹{invoice['net_amt']:,.2f}"
            ws.cell(row=current_row, column=11).value = f"₹{invoice['paid_amt']:,.2f}"
            ws.cell(row=current_row, column=12).value = f"₹{invoice['outstanding']:,.2f}"
            ws.cell(row=current_row, column=13).value = invoice['status']
            
            # Status color
            status_cell = ws.cell(row=current_row, column=13)
            if invoice['status'] == 'Paid':
                status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif invoice['status'] == 'Partial':
                status_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            # Borders
            for col in range(1, 14):
                ws.cell(row=current_row, column=col).border = self.border
            
            current_row += 1
        
        # Totals row (FIXED: Using ₹ symbol)
        totals = report_data['totals']
        ws.cell(row=current_row, column=1).value = "TOTALS"
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=8).value = f"₹{totals['invoice_amt']:,.2f}"
        ws.cell(row=current_row, column=8).font = Font(bold=True)
        ws.cell(row=current_row, column=9).value = f"₹{totals['tax_amt']:,.2f}"
        ws.cell(row=current_row, column=9).font = Font(bold=True)
        ws.cell(row=current_row, column=10).value = f"₹{totals['net_amt']:,.2f}"
        ws.cell(row=current_row, column=10).font = Font(bold=True)
        ws.cell(row=current_row, column=11).value = f"₹{totals['paid_amt']:,.2f}"
        ws.cell(row=current_row, column=11).font = Font(bold=True)
        ws.cell(row=current_row, column=12).value = f"₹{totals['outstanding']:,.2f}"
        ws.cell(row=current_row, column=12).font = Font(bold=True)
        
        for col in range(1, 14):
            ws.cell(row=current_row, column=col).border = self.border
            ws.cell(row=current_row, column=col).fill = PatternFill(
                start_color=self.secondary_color,
                end_color=self.secondary_color,
                fill_type="solid"
            )
        
        # Column widths
        column_widths = [10, 20, 12, 15, 12, 12, 25, 14, 12, 12, 12, 14, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Save
        company_name = self.branding["company_name"].replace(" ", "_")
        filename = f"{company_name}_AP_Invoice_Register_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        self.logger.info(f"Excel generated: {filepath}")
        return str(filepath)
    
    def _add_company_header(self, ws, start_row: int) -> int:
        """Add company header - FIXED: Centered logo and company name"""
        current_row = start_row
        logo_path = self.branding_manager.get_logo_path(self.user_id)
        
        if logo_path and logo_path.exists():
            # Logo at column B (centered)
            img = XLImage(str(logo_path))
            img.width = 120
            img.height = 60
            ws.add_image(img, f'B{current_row}')
            
            # Company name centered over table
            ws.merge_cells(f'E{current_row}:I{current_row}')
            ws[f'E{current_row}'].value = self.branding["company_name"]
            ws[f'E{current_row}'].font = Font(bold=True, size=18, color=self.primary_color)
            ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 3
        else:
            # No logo - center company name
            ws.merge_cells(f'A{current_row}:M{current_row}')
            ws[f'A{current_row}'].value = self.branding["company_name"]
            ws[f'A{current_row}'].font = Font(bold=True, size=20, color=self.primary_color)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
            current_row += 1
        
        # Separator line
        ws.merge_cells(f'A{current_row}:M{current_row}')
        ws[f'A{current_row}'].fill = PatternFill(
            start_color=self.accent_color,
            end_color=self.accent_color,
            fill_type="solid"
        )
        ws.row_dimensions[current_row].height = 3
        
        return current_row