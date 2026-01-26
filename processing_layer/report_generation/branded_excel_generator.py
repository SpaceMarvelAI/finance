"""
Professional Branded Excel Report Generator
Generates Excel reports with company branding
"""

from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from shared.branding.company_branding import CompanyBrandingManager
from shared.config.logging_config import get_logger


logger = get_logger(__name__)


class BrandedExcelGenerator:
    """
    Professional Branded Excel Generator
    
    Features:
    - Company logo in header
    - Company name with brand colors
    - Custom color themes
    - Professional styling
    - Dynamic data from database
    """
    
    def __init__(
        self,
        user_id: str,
        output_dir: str = "./output/reports"
    ):
        self.user_id = user_id
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Initialize branding manager (needed for logo path lookup)
        self.branding_manager = CompanyBrandingManager()
        
        # Try to load branding from database first
        from sqlalchemy import create_engine, text
        from sqlalchemy.orm import sessionmaker
        import os
        
        try:
            # Get database connection
            db_url = f"postgresql://{os.getenv('DB_USER', 'postgres')}:{os.getenv('DB_PASSWORD', 'postgres')}@{os.getenv('DB_HOST', 'localhost')}:{os.getenv('DB_PORT', '5432')}/{os.getenv('DB_NAME', 'financial_automation')}"
            engine = create_engine(db_url)
            Session = sessionmaker(bind=engine)
            db_session = Session()
            
            # Get user's company
            query = text("""
                SELECT c.name, c.logo_url, c.primary_color, c.secondary_color
                FROM users u
                JOIN companies c ON u.company_id = c.id
                WHERE u.id = :user_id
            """)
            result = db_session.execute(query, {"user_id": user_id}).first()
            
            if result:
                self.branding = {
                    "company_name": result[0] or "Company",
                    "logo_path": result[1],
                    "colors": {
                        "primary": result[2] or "#1976D2",
                        "secondary": result[3] or "#424242",
                        "accent": "#FFC107"  # Default accent color
                    }
                }
                print(f" Loaded branding from database: {self.branding['company_name']}")
            else:
                raise Exception("No user found")
                
            db_session.close()
            
        except Exception as e:
            print(f"Ã¢Å¡Â Ã¯Â¸Â  Could not load branding from database: {e}")
            # Fallback to file-based branding
            self.branding = self.branding_manager.get_branding(user_id)
            
            if not self.branding:
                # Use default branding if none configured
                print(f"Ã¢Å¡Â Ã¯Â¸Â  No custom branding for user {user_id}, using defaults")
                self.branding = {
                    "company_name": "Company",
                    "logo_path": None,
                    "colors": {
                        "primary": "#1976D2",
                        "secondary": "#424242",
                        "accent": "#FFC107"
                    }
                }
        
        # Get colors
        colors = self.branding["colors"]
        self.primary_color = colors["primary"].lstrip('#')
        self.secondary_color = colors["secondary"].lstrip('#')
        self.accent_color = colors.get("accent", "#FFC107").lstrip('#')
        
        # Styling
        self.header_font = Font(bold=True, size=11, color="FFFFFF")
        self.header_fill = PatternFill(
            start_color=self.primary_color,
            end_color=self.primary_color,
            fill_type="solid"
        )
        self.title_font = Font(bold=True, size=16, color=self.primary_color)
        self.company_font = Font(bold=True, size=14, color=self.primary_color)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.logger = logger
    
    def generate_dso_report(self, dso_data: Dict[str, Any]) -> str:
        """
        Generate DSO Analysis Report with company branding
        
        Args:
            dso_data: DSO calculation result from DSOAgent
            
        Returns:
            Path to generated Excel file
        """
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DSO Analysis"
        
        current_row = 1
        
        # Add company header
        current_row = self._add_company_header(ws, current_row)
        
        # Report title
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = "DAYS SALES OUTSTANDING (DSO) ANALYSIS"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='center')
        
        # Report metadata
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        meta_cell = ws[f'A{current_row}']
        meta_cell.value = f"Period: {dso_data.get('start_date', '')} to {dso_data.get('end_date', '')} | Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        meta_cell.alignment = Alignment(horizontal='center')
        meta_cell.font = Font(italic=True, size=10)
        
        # DSO Summary Section
        current_row += 2
        
        # Extract actual DSO data from the nested structure
        summary_data = dso_data
        if 'data' in dso_data:
            summary_data = dso_data['data']
        
        # Summary header
        ws.merge_cells(f'A{current_row}:F{current_row}')
        summary_header = ws[f'A{current_row}']
        summary_header.value = "DSO PERFORMANCE SUMMARY"
        summary_header.font = Font(bold=True, size=12, color=self.primary_color)
        summary_header.alignment = Alignment(horizontal='center')
        
        # Summary metrics
        current_row += 1
        
        # DSO Value with performance indicator
        ws[f'A{current_row}'] = f"DSO: {summary_data.get('dso', 0)} days"
        ws[f'A{current_row}'].font = Font(bold=True, size=14)
        
        # Performance indicator
        performance = summary_data.get('performance', 'Unknown')
        category = summary_data.get('category', 'unknown')
        
        ws[f'D{current_row}'] = f"Performance: {performance}"
        ws[f'D{current_row}'].font = Font(bold=True, size=12)
        
        # Color code the performance
        perf_cell = ws[f'D{current_row}']
        if category == 'success':
            perf_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        elif category == 'warning':
            perf_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        else:
            perf_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        current_row += 1
        
        # Key metrics
        metrics = [
            ("Total Sales", f"₹{summary_data.get('total_sales', 0):,.2f}"),
            ("Average AR", f"₹{summary_data.get('average_ar', 0):,.2f}"),
            ("Outstanding AR", f"₹{summary_data.get('outstanding_ar', 0):,.2f}"),
            ("Invoice Count", summary_data.get('invoice_count', 0)),
            ("Paid Invoices", summary_data.get('paid_invoices', 0)),
            ("Unpaid Invoices", summary_data.get('unpaid_invoices', 0))
        ]
        
        for i, (label, value) in enumerate(metrics):
            row_offset = current_row + i
            ws[f'A{row_offset}'] = label
            ws[f'A{row_offset}'].font = Font(bold=True)
            ws[f'B{row_offset}'] = value
            ws[f'B{row_offset}'].alignment = Alignment(horizontal='right')
        
        current_row += len(metrics) + 1
        
        # Collection Analysis Section
        ws.merge_cells(f'A{current_row}:F{current_row}')
        analysis_header = ws[f'A{current_row}']
        analysis_header.value = "COLLECTION ANALYSIS"
        analysis_header.font = Font(bold=True, size=12, color=self.primary_color)
        analysis_header.alignment = Alignment(horizontal='center')
        
        current_row += 1
        
        # Collection efficiency
        collection_efficiency = (summary_data.get('paid_invoices', 0) / summary_data.get('invoice_count', 1) * 100) if summary_data.get('invoice_count', 0) > 0 else 0
        
        ws[f'A{current_row}'] = "Collection Efficiency"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = f"{collection_efficiency:.1f}%"
        ws[f'B{current_row}'].alignment = Alignment(horizontal='right')
        
        current_row += 1
        
        # Average collection period
        ws[f'A{current_row}'] = "Average Collection Period"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = f"{summary_data.get('dso', 0):.1f} days"
        ws[f'B{current_row}'].alignment = Alignment(horizontal='right')
        
        current_row += 2
        
        # Recommendations Section
        ws.merge_cells(f'A{current_row}:F{current_row}')
        rec_header = ws[f'A{current_row}']
        rec_header.value = "COLLECTION RECOMMENDATIONS"
        rec_header.font = Font(bold=True, size=12, color=self.primary_color)
        rec_header.alignment = Alignment(horizontal='center')
        
        current_row += 1
        
        recommendations = self._get_dso_recommendations(summary_data)
        
        for i, rec in enumerate(recommendations):
            ws[f'A{current_row + i}'] = f"• {rec}"
            ws[f'A{current_row + i}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        current_row += len(recommendations) + 2
        
        # Invoice Details Section
        ws.merge_cells(f'A{current_row}:F{current_row}')
        details_header = ws[f'A{current_row}']
        details_header.value = "INVOICE DETAILS"
        details_header.font = Font(bold=True, size=12, color=self.primary_color)
        details_header.alignment = Alignment(horizontal='center')
        
        current_row += 2
        
        # Table headers
        headers = ["Invoice No.", "Customer", "Invoice Date", "Due Date", "Amount", "Outstanding", "Days Outstanding"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        
        current_row += 1
        
        # Note: Since DSO agent doesn't return individual invoice details,
        # we'll create a summary table or leave it for future enhancement
        ws.merge_cells(f'A{current_row}:G{current_row}')
        note_cell = ws[f'A{current_row}']
        note_cell.value = "Note: Detailed invoice breakdown available in AR Aging Report"
        note_cell.font = Font(italic=True, size=10)
        note_cell.alignment = Alignment(horizontal='center')
        
        # Column widths
        ws.column_dimensions['A'].width = 15  # Invoice No
        ws.column_dimensions['B'].width = 20  # Customer
        ws.column_dimensions['C'].width = 12  # Invoice Date
        ws.column_dimensions['D'].width = 12  # Due Date
        ws.column_dimensions['E'].width = 15  # Amount
        ws.column_dimensions['F'].width = 15  # Outstanding
        ws.column_dimensions['G'].width = 15  # Days Outstanding
        
        # Save
        company_name = self.branding["company_name"].replace(" ", "_")
        filename = f"{company_name}_DSO_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        self.logger.info(f"DSO Analysis Excel generated: {filepath}")
        return str(filepath)
    
    def _get_dso_recommendations(self, dso_data: Dict[str, Any]) -> list:
        """
        Generate collection recommendations based on DSO performance
        """
        dso = dso_data.get('dso', 0)
        performance = dso_data.get('performance', 'Unknown')
        collection_efficiency = (dso_data.get('paid_invoices', 0) / dso_data.get('invoice_count', 1) * 100) if dso_data.get('invoice_count', 0) > 0 else 0
        
        recommendations = []
        
        if performance == "Excellent":
            recommendations.append("DSO performance is excellent. Continue current collection practices.")
            recommendations.append("Consider offering early payment discounts to maintain good relationships.")
            recommendations.append("Review credit terms for potential optimization opportunities.")
        
        elif performance == "Good":
            recommendations.append("DSO performance is good but can be improved.")
            recommendations.append("Implement stricter follow-up procedures for invoices over 30 days.")
            recommendations.append("Review customer credit limits and payment terms.")
        
        elif performance == "Fair":
            recommendations.append("DSO performance needs attention. Implement immediate improvements.")
            recommendations.append("Strengthen collection follow-up process for overdue accounts.")
            recommendations.append("Review and tighten credit approval policies.")
            recommendations.append("Consider offering payment plans for large outstanding balances.")
        
        else:  # Needs Improvement
            recommendations.append("DSO performance is poor. Immediate action required.")
            recommendations.append("Implement aggressive collection strategies for overdue accounts.")
            recommendations.append("Review and potentially restrict credit terms for high-risk customers.")
            recommendations.append("Consider engaging collection agencies for severely delinquent accounts.")
            recommendations.append("Analyze root causes of slow collections and address systemic issues.")
        
        # Additional recommendations based on collection efficiency
        if collection_efficiency < 70:
            recommendations.append("Collection efficiency is low. Focus on improving payment collection rates.")
        
        if dso_data.get('unpaid_invoices', 0) > 0:
            recommendations.append("Prioritize collection efforts on unpaid invoices to improve cash flow.")
        
        return recommendations
    
    def generate_ap_invoice_register(self, report_data: Dict[str, Any]) -> str:
        """
        Generate AP Invoice Register with company branding
        
        Args:
            report_data: Structured report data from database
            
        Returns:
            Path to generated Excel file
        """
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AP Invoice Register"
        
        current_row = 1
        
        # Add company header
        current_row = self._add_company_header(ws, current_row)
        
        # Report title
        current_row += 1
        ws.merge_cells(f'A{current_row}:M{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = "AP INVOICE REGISTER REPORT"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='center')
        
        # Report metadata
        current_row += 1
        ws.merge_cells(f'A{current_row}:M{current_row}')
        meta_cell = ws[f'A{current_row}']
        meta_cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        meta_cell.alignment = Alignment(horizontal='center')
        meta_cell.font = Font(italic=True, size=10)
        
        # Summary section
        current_row += 2
        summary = report_data['summary']
        
        summary_fill = PatternFill(
            start_color=self.secondary_color,
            end_color=self.secondary_color,
            fill_type="solid"
        )
        
        ws[f'A{current_row}'] = f"Total Invoices: {summary['total_invoices']}"
        ws[f'A{current_row}'].fill = summary_fill
        ws[f'A{current_row}'].font = Font(bold=True)
        
        ws[f'D{current_row}'] = f"Paid: {summary['paid_count']}"
        ws[f'D{current_row}'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        ws[f'G{current_row}'] = f"Partial: {summary['partial_count']}"
        ws[f'G{current_row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        ws[f'J{current_row}'] = f"Unpaid: {summary['unpaid_count']}"
        ws[f'J{current_row}'].fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # Table headers
        current_row += 2
        headers = [
            "Trans ID", "Vendor Name", "Vendor ID", "Invoice No.",
            "Invoice Date", "Due Date", "Description",
            "Net Amt", "Tax Amt", "Sub Total", "Paid Amt",
            "Outstanding", "Status"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data rows
        current_row += 1
        for invoice in report_data['invoices']:
            ws.cell(row=current_row, column=1).value = invoice['trans_id']
            ws.cell(row=current_row, column=2).value = invoice['vendor_name']
            ws.cell(row=current_row, column=3).value = invoice['vendor_id']
            ws.cell(row=current_row, column=4).value = invoice['invoice_no']
            ws.cell(row=current_row, column=5).value = invoice['invoice_date']
            ws.cell(row=current_row, column=6).value = invoice['due_date']
            ws.cell(row=current_row, column=7).value = invoice['description']
            ws.cell(row=current_row, column=8).value = f"₹{invoice['net_amt']:,.2f}"
            ws.cell(row=current_row, column=9).value = f"₹{invoice['tax_amt']:,.2f}"
            ws.cell(row=current_row, column=10).value = f"₹{invoice['sub_total']:,.2f}"
            ws.cell(row=current_row, column=11).value = f"₹{invoice['paid_amt']:,.2f}"
            ws.cell(row=current_row, column=12).value = f"₹{invoice['outstanding']:,.2f}"
            ws.cell(row=current_row, column=13).value = invoice['status']
            
            # Status color coding
            status_cell = ws.cell(row=current_row, column=13)
            if invoice['status'] == 'Paid':
                status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif invoice['status'] == 'Partial':
                status_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            # Borders
            for col in range(1, 14):
                ws.cell(row=current_row, column=col).border = self.border
                ws.cell(row=current_row, column=col).alignment = Alignment(vertical='center')
            
            current_row += 1
        
        # Totals row
        totals = report_data.get('totals') or {}
        if not totals or not totals.get('invoice_amt'):
            summary = report_data.get('summary', {})
            totals = {
                'invoice_amt': summary.get('total_amount', 0),
                'tax_amt': 0,  # Will calculate from invoices
                'net_amt': summary.get('total_amount', 0),
                'paid_amt': summary.get('total_paid', 0),
                'outstanding': summary.get('total_outstanding', 0)
            }
        ws.cell(row=current_row, column=1).value = "TOTALS"
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        
        ws.cell(row=current_row, column=8).value = f"₹{totals['invoice_amt']:,.2f}"
        ws.cell(row=current_row, column=8).value = f"₹{totals['invoice_amt']:,.2f}"
        ws.cell(row=current_row, column=8).font = Font(bold=True)
        ws.cell(row=current_row, column=9).value = f"₹{totals['tax_amt']:,.2f}"
        ws.cell(row=current_row, column=9).font = Font(bold=True)
        ws.cell(row=current_row, column=10).value = f"₹{totals['net_amt']:,.2f}"
        ws.cell(row=current_row, column=10).font = Font(bold=True)
        ws.cell(row=current_row, column=11).value = f"₹{totals['paid_amt']:,.2f}"
        ws.cell(row=current_row, column=11).font = Font(bold=True)
        ws.cell(row=current_row, column=12).value = f"₹{totals['outstanding']:,.2f}"
        ws.cell(row=current_row, column=12).font = Font(bold=True)
        
        # Borders on totals
        for col in range(1, 14):
            ws.cell(row=current_row, column=col).border = self.border
            ws.cell(row=current_row, column=col).fill = PatternFill(
                start_color=self.secondary_color,
                end_color=self.secondary_color,
                fill_type="solid"
            )
        
        # Column widths
        column_widths = [10, 20, 12, 15, 12, 12, 25, 14, 12, 12, 12, 14, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Save
        company_name = self.branding["company_name"].replace(" ", "_")
        filename = f"{company_name}_AP_Invoice_Register_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        self.logger.info(f"Branded Excel generated: {filepath}")
        return str(filepath)
    
    def _add_company_header(self, ws, start_row: int) -> int:
        """
        Add company header with logo and name
        
        Returns:
            Next available row
        """
        
        current_row = start_row
        
        # Check if logo exists
        logo_path = None
        if self.branding.get('logo_path'):
            from pathlib import Path
            logo_url = self.branding['logo_path']
            
            # Convert /static/logos/xxx.png to ./data/logos/xxx.png
            if logo_url.startswith('/static/'):
                logo_path = logo_url.replace('/static/', './data/')
                
                # Check if file exists and add logo
                logo_file = Path(logo_path)
                if logo_file.exists():
                    img = XLImage(str(logo_path))
                    img.width = min(img.width, 120)
                    img.height = min(img.height, 60)
                    ws.add_image(img, f'B{current_row}')
                else:
                    print(f" Logo not found at {logo_path}")
            
            # Resize (smaller for better fit)
            img.width = min(img.width, 120)
            img.height = min(img.height, 60)
            
            # Position logo at column B
            ws.add_image(img, f'B{current_row}')
            
            # Company name centered over table
            ws.merge_cells(f'E{current_row}:I{current_row}')
            name_cell = ws[f'E{current_row}']
            name_cell.value = self.branding["company_name"]
            name_cell.font = Font(bold=True, size=18, color=self.primary_color)
            name_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Advance rows (reduced spacing)
            current_row += 3
        else:
            # No logo - just company name
            ws.merge_cells(f'A{current_row}:M{current_row}')
            name_cell = ws[f'A{current_row}']
            name_cell.value = self.branding["company_name"]
            name_cell.font = Font(bold=True, size=20, color=self.primary_color)
            name_cell.alignment = Alignment(horizontal='center')
            
            current_row += 1
        
        # Add a separator line
        ws.merge_cells(f'A{current_row}:M{current_row}')
        separator = ws[f'A{current_row}']
        separator.fill = PatternFill(
            start_color=self.accent_color,
            end_color=self.accent_color,
            fill_type="solid"
        )
        ws.row_dimensions[current_row].height = 3
        
        return current_row
    
    def generate_ap_aging(self, report_data: Dict[str, Any]) -> str:
        """
        Generate AP Aging Report with company branding
        
        Table Format (EXACT):
        Vendor Name | Invoice No. | Due Date | Total Due | Current (0-30) | 31-60 Days | 61-90 Days | 90+ Days
        """
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AP Aging"
        
        current_row = 1
        
        # Add company header
        current_row = self._add_company_header(ws, current_row)
        
        # Report title - support both AP and AR
        current_row += 1
        ws.merge_cells(f'A{current_row}:H{current_row}')
        title_cell = ws[f'A{current_row}']
        report_type = report_data.get('report_type', 'AP_AGING')
        if 'AR' in report_type.upper():
            title_cell.value = "AR AGING REPORT"
            ws.title = "AR Aging"
        else:
            title_cell.value = "AP AGING REPORT"
            ws.title = "AP Aging"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='center')
        
        # Report date
        current_row += 1
        ws.merge_cells(f'A{current_row}:H{current_row}')
        date_cell = ws[f'A{current_row}']
        date_cell.value = f"As of: {report_data['report_metadata']['as_of_date']} | Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        date_cell.alignment = Alignment(horizontal='center')
        date_cell.font = Font(size=9)
        
        current_row += 2
        
        # TABLE - Exact format from screenshot
        table_data = report_data.get('table_data') or report_data.get('aging_data')
        
        # Headers
        headers = table_data['headers']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        
        current_row += 1
        
        # Data rows (from database - NOT hardcoded!)
        for row_data in table_data['rows']:
            ws.cell(row=current_row, column=1).value = row_data['vendor_name']
            ws.cell(row=current_row, column=2).value = row_data['invoice_no']
            ws.cell(row=current_row, column=3).value = row_data['due_date']
            ws.cell(row=current_row, column=4).value = f"₹{row_data['total_due']:,.2f}"
            ws.cell(row=current_row, column=5).value = f"₹{row_data['current_0_30']:,.2f}"
            ws.cell(row=current_row, column=6).value = f"₹{row_data['days_31_60']:,.2f}"
            ws.cell(row=current_row, column=7).value = f"₹{row_data['days_61_90']:,.2f}"
            ws.cell(row=current_row, column=8).value = f"₹{row_data['days_90_plus']:,.2f}"
            
            # Alignment
            for col in range(1, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.border = self.border
                if col == 1:  # Vendor name - left align
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:  # Numbers - right align
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            
            current_row += 1
        
        # GRAND TOTALS row
        totals = table_data['grand_totals']
        
        ws.cell(row=current_row, column=1).value = "GRAND TOTALS"
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        ws.cell(row=current_row, column=4).value = f"₹{totals['total_due']:,.2f}"
        ws.cell(row=current_row, column=4).font = Font(bold=True, size=11)
        
        ws.cell(row=current_row, column=5).value = f"₹{totals['current_0_30']:,.2f}"
        ws.cell(row=current_row, column=5).font = Font(bold=True, size=11)
        
        ws.cell(row=current_row, column=6).value = f"₹{totals['days_31_60']:,.2f}"
        ws.cell(row=current_row, column=6).font = Font(bold=True, size=11)
        
        ws.cell(row=current_row, column=7).value = f"₹{totals['days_61_90']:,.2f}"
        ws.cell(row=current_row, column=7).font = Font(bold=True, size=11)
        
        ws.cell(row=current_row, column=8).value = f"₹{totals['days_90_plus']:,.2f}"
        ws.cell(row=current_row, column=8).font = Font(bold=True, size=11)
        
        # Grand totals styling
        for col in range(1, 9):
            cell = ws.cell(row=current_row, column=col)
            cell.border = self.border
            cell.fill = PatternFill(
                start_color=self.secondary_color,
                end_color=self.secondary_color,
                fill_type="solid"
            )
            if col > 1:
                cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Column widths
        ws.column_dimensions['A'].width = 20  # Vendor Name
        ws.column_dimensions['B'].width = 15  # Invoice No
        ws.column_dimensions['C'].width = 12  # Due Date
        ws.column_dimensions['D'].width = 15  # Total Due
        ws.column_dimensions['E'].width = 15  # Current (0-30)
        ws.column_dimensions['F'].width = 15  # 31-60 Days
        ws.column_dimensions['G'].width = 15  # 61-90 Days
        ws.column_dimensions['H'].width = 15  # 90+ Days
        
        # Save
        company_name = self.branding["company_name"].replace(" ", "_")
        report_type_str = report_data.get('report_type', 'AP_AGING')
        if 'AR' in report_type_str.upper():
            report_name = "AR_Aging"
            log_msg = "AR Aging Excel generated"
        else:
            report_name = "AP_Aging"
            log_msg = "AP Aging Excel generated"
        
        filename = f"{company_name}_{report_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        self.logger.info(f"{log_msg}: {filepath}")
        return str(filepath)
    
    def generate_ap_overdue(self, report_data: Dict[str, Any], sla_days: int = 30) -> str:
        """
        Generate AP Overdue SLA Report with company branding
        
        Table Format:
        Vendor Name | Invoice No. | Due Date | Total Due | Breach Days | SLA Severity | Status
        """
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AP Overdue SLA"
        
        current_row = 1
        
        # Add company header
        current_row = self._add_company_header(ws, current_row)
        
        # Report title
        current_row += 1
        ws.merge_cells(f'A{current_row}:G{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = "AP OVERDUE SLA REPORT"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='center')
        
        # Report metadata
        current_row += 1
        ws.merge_cells(f'A{current_row}:G{current_row}')
        meta_cell = ws[f'A{current_row}']
        meta_cell.value = f"SLA Days: {sla_days} | Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        meta_cell.alignment = Alignment(horizontal='center')
        meta_cell.font = Font(italic=True, size=10)
        
        # Summary section
        current_row += 2
        summary = report_data['summary']
        
        summary_fill = PatternFill(
            start_color=self.secondary_color,
            end_color=self.secondary_color,
            fill_type="solid"
        )
        
        ws[f'A{current_row}'] = f"Total Breaches: {summary['total_breached']}"
        ws[f'A{current_row}'].fill = summary_fill
        ws[f'A{current_row}'].font = Font(bold=True)
        
        ws[f'D{current_row}'] = f"Total Amount: ₹{summary['total_amount']:,.2f}"
        ws[f'D{current_row}'].fill = summary_fill
        ws[f'D{current_row}'].font = Font(bold=True)
        
        # Severity breakdown
        current_row += 1
        severity_text = " | ".join([f"{k}: {v}" for k, v in summary['by_severity'].items()])
        ws[f'A{current_row}'] = f"By Severity: {severity_text}"
        ws[f'A{current_row}'].font = Font(size=10)
        
        # Table headers
        current_row += 2
        headers = [
            "Vendor Name", "Invoice No.", "Due Date", "Total Due",
            "Breach Days", "SLA Severity", "Status"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        
        current_row += 1
        
        # Data rows
        for invoice in report_data['invoices']:
            ws.cell(row=current_row, column=1).value = invoice['vendor_name']
            ws.cell(row=current_row, column=2).value = invoice.get('invoice_no', invoice.get('invoice_number', ''))
            ws.cell(row=current_row, column=3).value = invoice['due_date']
            ws.cell(row=current_row, column=4).value = f"₹{invoice['outstanding']:,.2f}"
            ws.cell(row=current_row, column=5).value = invoice.get('breach_days', 0)
            ws.cell(row=current_row, column=6).value = invoice.get('sla_severity', 'Unknown')
            ws.cell(row=current_row, column=7).value = "OVERDUE"
            
            # Severity color coding
            severity_cell = ws.cell(row=current_row, column=6)
            severity = invoice.get('sla_severity', 'Unknown')
            if severity == 'Critical':
                severity_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                severity_cell.font = Font(color="FFFFFF", bold=True)
            elif severity == 'High':
                severity_cell.fill = PatternFill(start_color="FF4500", end_color="FF4500", fill_type="solid")
                severity_cell.font = Font(color="FFFFFF", bold=True)
            elif severity == 'Medium':
                severity_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                severity_cell.font = Font(color="FFFFFF", bold=True)
            elif severity == 'Low':
                severity_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                severity_cell.font = Font(color="000000", bold=True)
            
            # Status cell styling
            status_cell = ws.cell(row=current_row, column=7)
            status_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            status_cell.font = Font(bold=True)
            
            # Alignment
            for col in range(1, 8):
                cell = ws.cell(row=current_row, column=col)
                cell.border = self.border
                if col in [1, 2, 6, 7]:  # Text columns
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:  # Number columns
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            
            current_row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 20  # Vendor Name
        ws.column_dimensions['B'].width = 15  # Invoice No
        ws.column_dimensions['C'].width = 12  # Due Date
        ws.column_dimensions['D'].width = 15  # Total Due
        ws.column_dimensions['E'].width = 12  # Breach Days
        ws.column_dimensions['F'].width = 15  # SLA Severity
        ws.column_dimensions['G'].width = 12  # Status
        
        # Save
        company_name = self.branding["company_name"].replace(" ", "_")
        filename = f"{company_name}_AP_Overdue_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        self.logger.info(f"AP Overdue Excel generated: {filepath}")
        return str(filepath)
    def generate_ar_invoice_register(self, report_data: Dict[str, Any]) -> str:
        """
        Generate AR Invoice Register with company branding
        Identical to AP Register but for customer invoices
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AR Register"
        
        current_row = 1
        current_row = self._add_company_header(ws, current_row)
        
        # Report title
        current_row += 1
        ws.merge_cells(f'A{current_row}:M{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = "AR INVOICE REGISTER REPORT"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='center')
        
        # Date
        current_row += 1
        ws.merge_cells(f'A{current_row}:M{current_row}')
        date_cell = ws[f'A{current_row}']
        date_cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        date_cell.alignment = Alignment(horizontal='center')
        date_cell.font = Font(size=9)
        
        # Summary section
        current_row += 2
        summary = report_data['summary']
        
        summary_fill = PatternFill(
            start_color=self.secondary_color,
            end_color=self.secondary_color,
            fill_type="solid"
        )
        
        ws[f'A{current_row}'] = f"Total Invoices: {summary['total_invoices']}"
        ws[f'A{current_row}'].fill = summary_fill
        ws[f'A{current_row}'].font = Font(bold=True)
        
        ws[f'D{current_row}'] = f"Paid: {summary.get('paid_count', 0)}"
        ws[f'D{current_row}'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        ws[f'G{current_row}'] = f"Partial: {summary.get('partial_count', 0)}"
        ws[f'G{current_row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        ws[f'J{current_row}'] = f"Unpaid: {summary.get('unpaid_count', 0)}"
        ws[f'J{current_row}'].fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # Table headers
        current_row += 2
        headers = [
            "Trans ID", "Customer Name", "Customer ID", "Invoice No.",
            "Invoice Date", "Due Date", "Description",
            "Net Amt", "Tax Amt", "Sub Total", "Received Amt",
            "Outstanding", "Status"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data rows
        current_row += 1
        for invoice in report_data['invoices']:
            ws.cell(row=current_row, column=1).value = str(invoice.get('id', ''))[:8]
            ws.cell(row=current_row, column=2).value = invoice.get('customer_name', 'Unknown')
            ws.cell(row=current_row, column=3).value = str(invoice.get('customer_id', ''))[:12]
            ws.cell(row=current_row, column=4).value = invoice.get('invoice_number', '')
            ws.cell(row=current_row, column=5).value = invoice.get('invoice_date', '')
            ws.cell(row=current_row, column=6).value = invoice.get('due_date', '')
            ws.cell(row=current_row, column=7).value = invoice.get('description', 'General Invoice')
            ws.cell(row=current_row, column=8).value = f"₹{float(invoice.get('inr_amount', 0)):,.2f}"
            ws.cell(row=current_row, column=9).value = f"₹{float(invoice.get('tax_amount', 0)):,.2f}"
            ws.cell(row=current_row, column=10).value = f"₹{float(invoice.get('inr_amount', 0)) - float(invoice.get('tax_amount', 0)):,.2f}"
            ws.cell(row=current_row, column=11).value = f"₹{float(invoice.get('received_amount', 0)):,.2f}"
            ws.cell(row=current_row, column=12).value = f"₹{float(invoice.get('outstanding', 0)):,.2f}"
            ws.cell(row=current_row, column=13).value = invoice.get('status', 'Unpaid')
            
            # Status color coding
            status_cell = ws.cell(row=current_row, column=13)
            if invoice.get('status') == 'Paid':
                status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif invoice.get('status') == 'Partial':
                status_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            # Borders
            for col in range(1, 14):
                ws.cell(row=current_row, column=col).border = self.border
                ws.cell(row=current_row, column=col).alignment = Alignment(vertical='center')
            
            current_row += 1
        
        # Totals row
        totals = {
            'invoice_amt': summary.get('total_amount', 0),
            'tax_amt': sum(float(inv.get('tax_amount', 0)) for inv in report_data['invoices']),
            'net_amt': summary.get('total_amount', 0) - sum(float(inv.get('tax_amount', 0)) for inv in report_data['invoices']),
            'received_amt': summary.get('total_received', 0),
            'outstanding': summary.get('total_outstanding', 0)
        }
        
        ws.cell(row=current_row, column=1).value = "TOTALS"
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        
        ws.cell(row=current_row, column=8).value = f"₹{totals['invoice_amt']:,.2f}"
        ws.cell(row=current_row, column=8).font = Font(bold=True)
        ws.cell(row=current_row, column=9).value = f"₹{totals['tax_amt']:,.2f}"
        ws.cell(row=current_row, column=9).font = Font(bold=True)
        ws.cell(row=current_row, column=10).value = f"₹{totals['net_amt']:,.2f}"
        ws.cell(row=current_row, column=10).font = Font(bold=True)
        ws.cell(row=current_row, column=11).value = f"₹{totals['received_amt']:,.2f}"
        ws.cell(row=current_row, column=11).font = Font(bold=True)
        ws.cell(row=current_row, column=12).value = f"₹{totals['outstanding']:,.2f}"
        ws.cell(row=current_row, column=12).font = Font(bold=True)
        
        # Borders on totals
        for col in range(1, 14):
            ws.cell(row=current_row, column=col).border = self.border
            ws.cell(row=current_row, column=col).fill = PatternFill(
                start_color=self.secondary_color,
                end_color=self.secondary_color,
                fill_type="solid"
            )
        
        # Column widths
        column_widths = [10, 20, 12, 15, 12, 12, 25, 14, 12, 12, 12, 14, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Save
        company_name = self.branding["company_name"].replace(" ", "_")
        filename = f"{company_name}_AR_Register_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        self.logger.info(f"AR Register Excel generated: {filepath}")
        return str(filepath)

    def generate_ar_aging(self, report_data: Dict[str, Any]) -> str:
        """
        Generate AR Aging Report with company branding
        Identical to AP Aging but for customer invoices
        """
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AR Aging"
        
        current_row = 1
        
        # Add company header
        current_row = self._add_company_header(ws, current_row)
        
        # Report title - AR specific
        current_row += 1
        ws.merge_cells(f'A{current_row}:H{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = "AR AGING REPORT"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='center')
        
        # Report date
        current_row += 1
        ws.merge_cells(f'A{current_row}:H{current_row}')
        date_cell = ws[f'A{current_row}']
        date_cell.value = f"As of: {report_data['report_metadata']['as_of_date']} | Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        date_cell.alignment = Alignment(horizontal='center')
        date_cell.font = Font(size=9)
        
        current_row += 2
        
        # TABLE - Exact format from screenshot
        table_data = report_data.get('table_data') or report_data.get('aging_data')
        
        # Headers
        headers = table_data['headers']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        
        current_row += 1
        
        # Data rows (from database - NOT hardcoded!)
        for row_data in table_data['rows']:
            ws.cell(row=current_row, column=1).value = row_data['vendor_name']  # This should be customer_name for AR
            ws.cell(row=current_row, column=2).value = row_data['invoice_no']
            ws.cell(row=current_row, column=3).value = row_data['due_date']
            ws.cell(row=current_row, column=4).value = f"₹{row_data['total_due']:,.2f}"
            ws.cell(row=current_row, column=5).value = f"₹{row_data['current_0_30']:,.2f}"
            ws.cell(row=current_row, column=6).value = f"₹{row_data['days_31_60']:,.2f}"
            ws.cell(row=current_row, column=7).value = f"₹{row_data['days_61_90']:,.2f}"
            ws.cell(row=current_row, column=8).value = f"₹{row_data['days_90_plus']:,.2f}"
            
            # Alignment
            for col in range(1, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.border = self.border
                if col == 1:  # Customer name - left align
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:  # Numbers - right align
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            
            current_row += 1
        
        # GRAND TOTALS row
        totals = table_data['grand_totals']
        
        ws.cell(row=current_row, column=1).value = "GRAND TOTALS"
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        ws.cell(row=current_row, column=4).value = f"₹{totals['total_due']:,.2f}"
        ws.cell(row=current_row, column=4).font = Font(bold=True, size=11)
        
        ws.cell(row=current_row, column=5).value = f"₹{totals['current_0_30']:,.2f}"
        ws.cell(row=current_row, column=5).font = Font(bold=True, size=11)
        
        ws.cell(row=current_row, column=6).value = f"₹{totals['days_31_60']:,.2f}"
        ws.cell(row=current_row, column=6).font = Font(bold=True, size=11)
        
        ws.cell(row=current_row, column=7).value = f"₹{totals['days_61_90']:,.2f}"
        ws.cell(row=current_row, column=7).font = Font(bold=True, size=11)
        
        ws.cell(row=current_row, column=8).value = f"₹{totals['days_90_plus']:,.2f}"
        ws.cell(row=current_row, column=8).font = Font(bold=True, size=11)
        
        # Grand totals styling
        for col in range(1, 9):
            cell = ws.cell(row=current_row, column=col)
            cell.border = self.border
            cell.fill = PatternFill(
                start_color=self.secondary_color,
                end_color=self.secondary_color,
                fill_type="solid"
            )
            if col > 1:
                cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Column widths
        ws.column_dimensions['A'].width = 20  # Customer Name
        ws.column_dimensions['B'].width = 15  # Invoice No
        ws.column_dimensions['C'].width = 12  # Due Date
        ws.column_dimensions['D'].width = 15  # Total Due
        ws.column_dimensions['E'].width = 15  # Current (0-30)
        ws.column_dimensions['F'].width = 15  # 31-60 Days
        ws.column_dimensions['G'].width = 15  # 61-90 Days
        ws.column_dimensions['H'].width = 15  # 90+ Days
        
        # Save
        company_name = self.branding["company_name"].replace(" ", "_")
        filename = f"{company_name}_AR_Aging_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        self.logger.info(f"AR Aging Excel generated: {filepath}")
        return str(filepath)

    def generate_ar_collection(self, report_data: Dict[str, Any]) -> str:
        """
        Generate AR Collection Priority Report with company branding
        
        Table Format:
        Customer Name | Invoice No. | Due Date | Outstanding | Priority Score | SLA Breach | Action
        """
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AR Collection"
        
        current_row = 1
        
        # Add company header
        current_row = self._add_company_header(ws, current_row)
        
        # Report title
        current_row += 1
        ws.merge_cells(f'A{current_row}:G{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = "AR COLLECTION PRIORITY REPORT"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='center')
        
        # Report metadata
        current_row += 1
        ws.merge_cells(f'A{current_row}:G{current_row}')
        meta_cell = ws[f'A{current_row}']
        sla_days = report_data.get('sla_days', 30)
        meta_cell.value = f"SLA Days: {sla_days} | Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        meta_cell.alignment = Alignment(horizontal='center')
        meta_cell.font = Font(italic=True, size=10)
        
        # Summary section
        current_row += 2
        summary = report_data['summary']
        
        summary_fill = PatternFill(
            start_color=self.secondary_color,
            end_color=self.secondary_color,
            fill_type="solid"
        )
        
        ws[f'A{current_row}'] = f"Total Overdue: {summary['total_overdue']}"
        ws[f'A{current_row}'].fill = summary_fill
        ws[f'A{current_row}'].font = Font(bold=True)
        
        ws[f'D{current_row}'] = f"Total Outstanding: ₹{summary['total_outstanding']:,.2f}"
        ws[f'D{current_row}'].fill = summary_fill
        ws[f'D{current_row}'].font = Font(bold=True)
        
        # Priority breakdown
        current_row += 1
        priority_text = " | ".join([f"{k}: {v}" for k, v in summary['by_priority'].items()])
        ws[f'A{current_row}'] = f"By Priority: {priority_text}"
        ws[f'A{current_row}'].font = Font(size=10)
        
        # Table headers
        current_row += 2
        headers = [
            "Customer Name", "Invoice No.", "Due Date", "Outstanding",
            "Priority Score", "SLA Breach", "Action"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        
        current_row += 1
        
        # Data rows
        for invoice in report_data['invoices']:
            ws.cell(row=current_row, column=1).value = invoice.get('customer_name', 'Unknown')
            ws.cell(row=current_row, column=2).value = invoice.get('invoice_number', '')
            ws.cell(row=current_row, column=3).value = invoice.get('due_date', '')
            ws.cell(row=current_row, column=4).value = f"₹{float(invoice.get('outstanding', 0)):,.2f}"
            ws.cell(row=current_row, column=5).value = invoice.get('priority_score', 0)
            ws.cell(row=current_row, column=6).value = "YES" if invoice.get('sla_breach', False) else "NO"
            ws.cell(row=current_row, column=7).value = invoice.get('action', 'Follow up')
            
            # Priority color coding
            priority_cell = ws.cell(row=current_row, column=5)
            priority_score = invoice.get('priority_score', 0)
            if priority_score > 1000000:  # High priority
                priority_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                priority_cell.font = Font(color="FFFFFF", bold=True)
            elif priority_score > 500000:  # Medium priority
                priority_cell.fill = PatternFill(start_color="FF4500", end_color="FF4500", fill_type="solid")
                priority_cell.font = Font(color="FFFFFF", bold=True)
            elif priority_score > 100000:  # Low priority
                priority_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                priority_cell.font = Font(color="FFFFFF", bold=True)
            
            # SLA Breach cell styling
            sla_cell = ws.cell(row=current_row, column=6)
            if invoice.get('sla_breach', False):
                sla_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                sla_cell.font = Font(bold=True)
            
            # Alignment
            for col in range(1, 8):
                cell = ws.cell(row=current_row, column=col)
                cell.border = self.border
                if col in [1, 2, 7]:  # Text columns
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:  # Number columns
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            
            current_row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 20  # Customer Name
        ws.column_dimensions['B'].width = 15  # Invoice No
        ws.column_dimensions['C'].width = 12  # Due Date
        ws.column_dimensions['D'].width = 15  # Outstanding
        ws.column_dimensions['E'].width = 15  # Priority Score
        ws.column_dimensions['F'].width = 12  # SLA Breach
        ws.column_dimensions['G'].width = 30  # Action
        
        # Save
        company_name = self.branding["company_name"].replace(" ", "_")
        filename = f"{company_name}_AR_Collection_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        self.logger.info(f"AR Collection Excel generated: {filepath}")
        return str(filepath)
