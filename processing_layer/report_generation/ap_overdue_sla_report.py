"""
AP OVERDUE & SLA BREACH REPORT
Identifies invoices that have breached payment SLA

SLA Breach = (Current Date - Due Date) > SLA Days
"""

from typing import Dict, Any, Optional, List
from datetime import datetime, date, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from shared.config.logging_config import get_logger


logger = get_logger(__name__)


class APOverdueSLAGenerator:
    """
    AP Overdue & SLA Breach Report
    
    Shows:
    - Invoices past due date
    - SLA breach status
    - Overdue days
    - Priority escalation
    """
    
    def __init__(self, mcp_tools=None, sla_days: int = 30):
        """
        Initialize generator
        
        Args:
            mcp_tools: MCP tools instance
            sla_days: SLA threshold in days (default: 30)
        """
        self.mcp_tools = mcp_tools
        self.sla_days = sla_days
        
        from data_layer.database.database_manager import get_database
        self.db = get_database()
        
        self.logger = logger
        self.logger.info(f"APOverdueSLAGenerator initialized (SLA: {sla_days} days)")
    
    def generate_report(
        self, 
        company_id: Optional[str] = None,
        as_of_date: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Generate SLA Breach Report
        
        Args:
            company_id: Filter by company
            as_of_date: Date for calculation (default: today)
            
        Returns:
            Report data with SLA breaches
        """
        
        # Parse date
        if as_of_date:
            try:
                current_date = datetime.fromisoformat(as_of_date.replace('Z', '+00:00')).date()
            except:
                current_date = date.today()
        else:
            current_date = date.today()
        
        # Get unpaid purchase invoices from database
        try:
            all_docs = self.db.get_documents_by_category('purchase', company_id)
            self.logger.info(f"Retrieved {len(all_docs)} purchase documents")
        except Exception as e:
            self.logger.error(f"Database error: {e}")
            all_docs = []
        
        overdue_invoices = []
        sla_breached = []
        within_sla = []
        
        for doc_entry in all_docs:
            # Get amounts (INR)
            invoice_amount = float(doc_entry.get("grand_total") or 0)  # FIX: Use grand_total instead of inr_amount
            paid_amount = float(doc_entry.get("paid_amount") or 0)
            
            # Calculate outstanding
            outstanding = invoice_amount - paid_amount
            
            # Skip if paid
            if outstanding <= 0:
                continue
            
            # Get fields
            vendor_name = doc_entry.get("vendor_name") or "Unknown"
            invoice_no = doc_entry.get("document_number") or "N/A"
            invoice_date = doc_entry.get("document_date")
            
            # Use invoice date as due date (fallback)
            due_date_str = doc_entry.get("document_date")
            due_date = self._parse_date(due_date_str)
            
            if not due_date:
                continue
            
            # Calculate overdue days
            overdue_days = (current_date - due_date).days
            
            # Skip if not overdue
            if overdue_days <= 0:
                continue
            
            # Determine SLA status
            if overdue_days > self.sla_days:
                sla_status = "SLA Breached"
                priority = "High"
                sla_breached.append(invoice_no)
            else:
                sla_status = "Within SLA"
                priority = "Medium"
                within_sla.append(invoice_no)
            
            # Create row
            row = {
                "invoice_no": invoice_no,
                "vendor_name": vendor_name,
                "invoice_date": due_date.strftime("%m/%d/%Y"),
                "due_date": due_date.strftime("%m/%d/%Y"),
                "overdue_days": overdue_days,
                "outstanding": round(outstanding, 2),
                "sla_status": sla_status,
                "priority": priority,
                "sla_threshold": self.sla_days
            }
            
            overdue_invoices.append(row)
        
        # Sort by overdue days (most overdue first)
        overdue_invoices.sort(key=lambda x: x['overdue_days'], reverse=True)
        
        total_overdue_amount = sum(inv['outstanding'] for inv in overdue_invoices)
        breached_amount = sum(inv['outstanding'] for inv in overdue_invoices if inv['sla_status'] == "SLA Breached")
        
        self.logger.info(f"Found {len(overdue_invoices)} overdue invoices, {len(sla_breached)} SLA breached")
        
        return {
            "report_metadata": {
                "report_type": "AP_OVERDUE_SLA",
                "report_name": "AP Overdue & SLA Breach Report",
                "as_of_date": current_date.isoformat(),
                "generated_at": datetime.now().isoformat(),
                "sla_days": self.sla_days,
                "currency": "INR"
            },
            "summary": {
                "total_overdue": len(overdue_invoices),
                "sla_breached_count": len(sla_breached),
                "within_sla_count": len(within_sla),
                "total_overdue_amount": round(total_overdue_amount, 2),
                "breached_amount": round(breached_amount, 2),
                "breach_rate": round((len(sla_breached) / len(overdue_invoices) * 100) if overdue_invoices else 0, 1)
            },
            "overdue_data": {
                "invoices": overdue_invoices,
                "sla_breached": sla_breached,
                "within_sla": within_sla
            }
        }
    
    def _parse_date(self, date_str: str) -> Optional[date]:
        """Parse date string"""
        if not date_str:
            return None
        
        try:
            dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
            return dt.date()
        except:
            try:
                for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"]:
                    try:
                        dt = datetime.strptime(date_str, fmt)
                        return dt.date()
                    except:
                        continue
            except:
                pass
        
        return None


# =============================================================================
# EXCEL GENERATOR
# =============================================================================

class APOverdueSLAExcelGenerator:
    """Generate branded Excel for SLA report"""
    
    def __init__(self, user_id: str, output_dir: str = "./output/reports"):
        self.user_id = user_id
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        from shared.branding.company_branding import CompanyBrandingManager
        self.branding_manager = CompanyBrandingManager()
        self.branding = self.branding_manager.get_branding(user_id)
        
        if not self.branding:
            raise ValueError(f"No branding for user: {user_id}")
        
        colors = self.branding["colors"]
        self.primary_color = colors["primary"].lstrip('#')
        self.secondary_color = colors["secondary"].lstrip('#')
        self.accent_color = colors["accent"].lstrip('#')
        
        self.header_font = Font(bold=True, size=11, color="FFFFFF")
        self.header_fill = PatternFill(
            start_color=self.primary_color,
            end_color=self.primary_color,
            fill_type="solid"
        )
        self.title_font = Font(bold=True, size=16, color=self.primary_color)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.logger = logger
    
    def generate(self, report_data: Dict[str, Any]) -> str:
        """Generate Excel file"""
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AP Overdue SLA"
        
        current_row = 1
        
        # Company header
        current_row = self._add_company_header(ws, current_row)
        
        # Title
        current_row += 1
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'].value = "AP OVERDUE & SLA BREACH REPORT"
        ws[f'A{current_row}'].font = self.title_font
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        
        current_row += 1
        metadata = report_data['report_metadata']
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'].value = f"As of: {metadata['as_of_date']} | SLA Threshold: {metadata['sla_days']} days"
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        ws[f'A{current_row}'].font = Font(size=9)
        
        # Summary
        current_row += 2
        summary = report_data['summary']
        ws[f'A{current_row}'] = f"Total Overdue: {summary['total_overdue']}"
        ws[f'D{current_row}'] = f"SLA Breached: {summary['sla_breached_count']}"
        ws[f'F{current_row}'] = f"Breach Rate: {summary['breach_rate']}%"
        ws[f'H{current_row}'] = f"Total: ₹{summary['total_overdue_amount']:,.2f}"
        
        # Color code summary
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'D{current_row}'].font = Font(bold=True, color="FF0000")  # Red
        ws[f'F{current_row}'].font = Font(bold=True)
        ws[f'H{current_row}'].font = Font(bold=True)
        
        # Headers
        current_row += 2
        headers = [
            "Invoice No", "Vendor", "Due Date", "Overdue Days",
            "Outstanding", "SLA Status", "Priority", "Action"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        current_row += 1
        
        # Data rows
        for invoice in report_data['overdue_data']['invoices']:
            ws.cell(row=current_row, column=1).value = invoice['invoice_no']
            ws.cell(row=current_row, column=2).value = invoice['vendor_name']
            ws.cell(row=current_row, column=3).value = invoice['due_date']
            ws.cell(row=current_row, column=4).value = invoice['overdue_days']
            ws.cell(row=current_row, column=5).value = f"₹{invoice['outstanding']:,.2f}"
            ws.cell(row=current_row, column=6).value = invoice['sla_status']
            ws.cell(row=current_row, column=7).value = invoice['priority']
            
            # Action based on status
            if invoice['sla_status'] == "SLA Breached":
                ws.cell(row=current_row, column=8).value = "URGENT - Escalate"
                ws.cell(row=current_row, column=6).fill = PatternFill(
                    start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"
                )
                ws.cell(row=current_row, column=7).fill = PatternFill(
                    start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"
                )
            else:
                ws.cell(row=current_row, column=8).value = "Follow up"
                ws.cell(row=current_row, column=6).fill = PatternFill(
                    start_color="FFD93D", end_color="FFD93D", fill_type="solid"
                )
            
            # Borders
            for col in range(1, 9):
                ws.cell(row=current_row, column=col).border = self.border
            
            current_row += 1
        
        # Column widths
        column_widths = [15, 25, 12, 12, 15, 15, 12, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Save
        company_name = self.branding["company_name"].replace(" ", "_")
        filename = f"{company_name}_AP_Overdue_SLA_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        self.logger.info(f"SLA report generated: {filepath}")
        return str(filepath)
    
    def _add_company_header(self, ws, start_row: int) -> int:
        """Add company header"""
        current_row = start_row
        logo_path = self.branding_manager.get_logo_path(self.user_id)
        
        if logo_path and logo_path.exists():
            img = XLImage(str(logo_path))
            img.width = 120
            img.height = 60
            ws.add_image(img, f'B{current_row}')
            
            ws.merge_cells(f'D{current_row}:F{current_row}')
            ws[f'D{current_row}'].value = self.branding["company_name"]
            ws[f'D{current_row}'].font = Font(bold=True, size=18, color=self.primary_color)
            ws[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 3
        else:
            ws.merge_cells(f'A{current_row}:H{current_row}')
            ws[f'A{current_row}'].value = self.branding["company_name"]
            ws[f'A{current_row}'].font = Font(bold=True, size=20, color=self.primary_color)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
            current_row += 1
        
        # Separator
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'].fill = PatternFill(
            start_color=self.accent_color,
            end_color=self.accent_color,
            fill_type="solid"
        )
        ws.row_dimensions[current_row].height = 3
        
        return current_row