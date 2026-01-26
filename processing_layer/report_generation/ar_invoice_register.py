"""
AR INVOICE REGISTER REPORT
Accounts Receivable - Customer invoices (Sales)

Shows all customer invoices issued by the company
"""

from typing import Dict, Any, Optional, List
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from shared.config.logging_config import get_logger


logger = get_logger(__name__)


class ARInvoiceRegisterGenerator:
    """
    AR Invoice Register Report
    
    Shows customer invoices (money owed TO us)
    """
    
    def __init__(self, mcp_tools=None):
        """Initialize with database access"""
        self.mcp_tools = mcp_tools
        
        from data_layer.database.database_manager import get_database
        self.db = get_database()
        
        self.logger = logger
        self.logger.info("ARInvoiceRegisterGenerator initialized")
    
    def generate_report(self, company_id: Optional[str] = None) -> Dict[str, Any]:
        """
        Generate AR Invoice Register
        
        Args:
            company_id: Optional company filter
            
        Returns:
            Report data with customer invoices
        """
        
        # Get SALES invoices from database
        try:
            all_docs = self.db.get_documents_by_category('sales', company_id)
            self.logger.info(f"Retrieved {len(all_docs)} sales documents")
        except Exception as e:
            self.logger.error(f"Database error: {e}")
            all_docs = []
        
        invoices = []
        totals = {
            "invoice_amt": 0.0,
            "received_amt": 0.0,
            "outstanding": 0.0
        }
        
        trans_id = 1001
        
        for doc_entry in all_docs:
            # Get amounts (INR)
            invoice_amt = float(doc_entry.get("inr_amount") or 0)
            received_amt = float(doc_entry.get("paid_amount") or 0)
            
            # Get customer info
            customer_name = doc_entry.get("customer_name") or doc_entry.get("buyer_name") or "Unknown"
            invoice_no = doc_entry.get("document_number") or "N/A"
            invoice_date = doc_entry.get("document_date") or ""
            due_date = doc_entry.get("due_date") or invoice_date
            
            # Get original currency info
            original_currency = doc_entry.get("original_currency") or "INR"
            original_amount = float(doc_entry.get("original_amount") or 0)
            
            # Description
            if original_currency != "INR":
                description = f"{original_currency} {original_amount:.2f}"
            else:
                description = "Sales Invoice"
            
            # Calculate outstanding
            outstanding = invoice_amt - received_amt
            
            # Determine status
            if received_amt >= invoice_amt:
                status = "Closed"
            elif received_amt > 0:
                status = "Partially Paid"
            else:
                status = "Open"
            
            # Format dates
            inv_date_str = self._format_date(invoice_date)
            due_date_str = self._format_date(due_date)
            
            # Create row
            invoice_row = {
                "trans_id": trans_id,
                "customer_name": customer_name,
                "invoice_no": invoice_no,
                "invoice_date": inv_date_str,
                "due_date": due_date_str,
                "description": description,
                "invoice_amt": round(invoice_amt, 2),
                "received_amt": round(received_amt, 2),
                "outstanding": round(outstanding, 2),
                "status": status
            }
            
            invoices.append(invoice_row)
            
            # Update totals
            totals["invoice_amt"] += invoice_amt
            totals["received_amt"] += received_amt
            totals["outstanding"] += outstanding
            
            trans_id += 1
        
        # Round totals
        for key in totals:
            totals[key] = round(totals[key], 2)
        
        self.logger.info(f"Generated AR report: {len(invoices)} invoices, total: ₹{totals['invoice_amt']:,.2f}")
        
        return {
            "report_metadata": {
                "report_type": "AR_INVOICE_REGISTER",
                "report_name": "AR Invoice Register Report",
                "generated_at": datetime.now().isoformat(),
                "currency": "INR"
            },
            "summary": {
                "total_invoices": len(invoices),
                "closed_count": sum(1 for inv in invoices if inv["status"] == "Closed"),
                "partial_count": sum(1 for inv in invoices if inv["status"] == "Partially Paid"),
                "open_count": sum(1 for inv in invoices if inv["status"] == "Open")
            },
            "invoices": invoices,
            "totals": totals
        }
    
    def _format_date(self, date_str: str) -> str:
        """Format date to MM/DD/YY"""
        if not date_str:
            return ""
        try:
            dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
            return dt.strftime("%m/%d/%y")
        except:
            return str(date_str)[:10]


# =============================================================================
# EXCEL GENERATOR
# =============================================================================

class ARInvoiceRegisterExcelGenerator:
    """Generate branded Excel for AR Invoice Register"""
    
    def __init__(self, user_id: str, output_dir: str = "./output/reports"):
        self.user_id = user_id
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        from shared.tools.user_settings import get_settings_manager
        settings_mgr = get_settings_manager()
        branding = settings_mgr.get_branding(user_id)
        
        if not branding:
            raise ValueError(f"No branding for user: {user_id}")
        
        self.branding = branding
        colors = branding["colors"]
        self.primary_color = colors["primary"].lstrip('#')
        self.secondary_color = colors["secondary"].lstrip('#')
        self.accent_color = colors["accent"].lstrip('#')
        
        self.header_font = Font(bold=True, size=11, color="FFFFFF")
        self.header_fill = PatternFill(
            start_color=self.primary_color,
            end_color=self.primary_color,
            fill_type="solid"
        )
        self.title_font = Font(bold=True, size=16, color=self.primary_color)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.logger = logger
    
    def generate(self, report_data: Dict[str, Any]) -> str:
        """Generate Excel file"""
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AR Invoice Register"
        
        current_row = 1
        
        # Company header
        current_row = self._add_company_header(ws, current_row)
        
        # Title
        current_row += 1
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws[f'A{current_row}'].value = "AR INVOICE REGISTER REPORT"
        ws[f'A{current_row}'].font = self.title_font
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws[f'A{current_row}'].value = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        ws[f'A{current_row}'].font = Font(size=9)
        
        # Summary
        current_row += 2
        summary = report_data['summary']
        ws[f'A{current_row}'] = f"Total Invoices: {summary['total_invoices']}"
        ws[f'D{current_row}'] = f"Closed: {summary['closed_count']}"
        ws[f'F{current_row}'] = f"Partial: {summary['partial_count']}"
        ws[f'H{current_row}'] = f"Open: {summary['open_count']}"
        
        # Headers
        current_row += 2
        headers = [
            "Trans ID", "Customer", "Invoice No.", "Invoice Date",
            "Due Date", "Description", "Invoice Amt", "Received Amt",
            "Outstanding", "Status"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        current_row += 1
        
        # Data rows
        for invoice in report_data['invoices']:
            ws.cell(row=current_row, column=1).value = invoice['trans_id']
            ws.cell(row=current_row, column=2).value = invoice['customer_name']
            ws.cell(row=current_row, column=3).value = invoice['invoice_no']
            ws.cell(row=current_row, column=4).value = invoice['invoice_date']
            ws.cell(row=current_row, column=5).value = invoice['due_date']
            ws.cell(row=current_row, column=6).value = invoice['description']
            ws.cell(row=current_row, column=7).value = f"₹{invoice['invoice_amt']:,.2f}"
            ws.cell(row=current_row, column=8).value = f"₹{invoice['received_amt']:,.2f}"
            ws.cell(row=current_row, column=9).value = f"₹{invoice['outstanding']:,.2f}"
            ws.cell(row=current_row, column=10).value = invoice['status']
            
            # Status color
            status_cell = ws.cell(row=current_row, column=10)
            if invoice['status'] == 'Closed':
                status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif invoice['status'] == 'Partially Paid':
                status_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            # Borders
            for col in range(1, 11):
                ws.cell(row=current_row, column=col).border = self.border
            
            current_row += 1
        
        # Totals row
        totals = report_data['totals']
        ws.cell(row=current_row, column=1).value = "TOTALS"
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=7).value = f"₹{totals['invoice_amt']:,.2f}"
        ws.cell(row=current_row, column=7).font = Font(bold=True)
        ws.cell(row=current_row, column=8).value = f"₹{totals['received_amt']:,.2f}"
        ws.cell(row=current_row, column=8).font = Font(bold=True)
        ws.cell(row=current_row, column=9).value = f"₹{totals['outstanding']:,.2f}"
        ws.cell(row=current_row, column=9).font = Font(bold=True)
        
        for col in range(1, 11):
            ws.cell(row=current_row, column=col).border = self.border
            ws.cell(row=current_row, column=col).fill = PatternFill(
                start_color=self.secondary_color,
                end_color=self.secondary_color,
                fill_type="solid"
            )
        
        # Column widths
        column_widths = [10, 20, 15, 12, 12, 20, 14, 14, 14, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Save
        company_name = self.branding["company_name"].replace(" ", "_")
        filename = f"{company_name}_AR_Invoice_Register_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        self.logger.info(f"AR Excel generated: {filepath}")
        return str(filepath)
    
    def _add_company_header(self, ws, start_row: int) -> int:
        """Add company header"""
        current_row = start_row
        
        from shared.tools.user_settings import get_settings_manager
        settings_mgr = get_settings_manager()
        settings = settings_mgr.get_user_settings(self.user_id)
        logo_path = settings.get('logo_path') if settings else None
        
        if logo_path and Path(logo_path).exists():
            img = XLImage(str(logo_path))
            img.width = 120
            img.height = 60
            ws.add_image(img, f'B{current_row}')
            
            ws.merge_cells(f'E{current_row}:H{current_row}')
            ws[f'E{current_row}'].value = self.branding["company_name"]
            ws[f'E{current_row}'].font = Font(bold=True, size=18, color=self.primary_color)
            ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 3
        else:
            ws.merge_cells(f'A{current_row}:J{current_row}')
            ws[f'A{current_row}'].value = self.branding["company_name"]
            ws[f'A{current_row}'].font = Font(bold=True, size=20, color=self.primary_color)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
            current_row += 1
        
        # Separator
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws[f'A{current_row}'].fill = PatternFill(
            start_color=self.accent_color,
            end_color=self.accent_color,
            fill_type="solid"
        )
        ws.row_dimensions[current_row].height = 3
        
        return current_row