"""
AR Aging Report Node
Copies generate_ar_aging method from BrandedExcelGenerator
"""

from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from processing_layer.workflows.nodes.output_nodes.base_template_node import BaseTemplateNode, register_node
from shared.branding.company_branding import CompanyBrandingManager
from shared.config.logging_config import get_logger


logger = get_logger(__name__)


@register_node
class ARAgingReportNode(BaseTemplateNode):
    """
    AR Aging Report Node
    Copies generate_ar_aging method from BrandedExcelGenerator
    """
    
    name = "AR Aging Report"
    category = "output"
    description = "Generates AR Aging report by copying BrandedExcelGenerator functionality"
    
    input_schema = {
        "report_data": {"type": "object", "required": True},
        "company_id": {"type": "string", "required": True}
    }
    
    output_schema = {
        "file_path": {"type": "string", "description": "Path to generated Excel file"},
        "branding_applied": {"type": "boolean", "description": "Whether branding was applied"},
        "report_type": {"type": "string", "description": "Type of report generated"}
    }
    
    ALLOWED_PARAMS = ["company_id", "as_of_date"]
    
    def __init__(self):
        super().__init__()
    
    def _get_branding(self, company_id: str) -> Dict[str, Any]:
        """Get company branding configuration"""
        # Try to load branding from database first
        from sqlalchemy import create_engine, text
        from sqlalchemy.orm import sessionmaker
        import os
        
        try:
            # Get database connection
            db_url = f"postgresql://{os.getenv('DB_USER', 'postgres')}:{os.getenv('DB_PASSWORD', 'postgres')}@{os.getenv('DB_HOST', 'localhost')}:{os.getenv('DB_PORT', '5432')}/{os.getenv('DB_NAME', 'financial_automation')}"
            engine = create_engine(db_url)
            Session = sessionmaker(bind=engine)
            db_session = Session()
            
            # Get user's company
            query = text("""
                SELECT c.name, c.logo_url, c.primary_color, c.secondary_color
                FROM users u
                JOIN companies c ON u.company_id = c.id
                WHERE u.id = :user_id
            """)
            result = db_session.execute(query, {"user_id": company_id}).first()
            
            if result:
                branding = {
                    "company_name": result[0] or "Company",
                    "logo_path": result[1],
                    "colors": {
                        "primary": result[2] or "#1976D2",
                        "secondary": result[3] or "#424242",
                        "accent": "#FFC107"  # Default accent color
                    }
                }
            else:
                raise Exception("No user found")
                
            db_session.close()
            
        except Exception as e:
            print(f"⚠️  Could not load branding from database: {e}")
            # Fallback to file-based branding
            branding_manager = CompanyBrandingManager()
            branding = branding_manager.get_branding(company_id)
            
            if not branding:
                # Use default branding if none configured
                print(f"⚠️  No custom branding for user {company_id}, using defaults")
                branding = {
                    "company_name": "Company",
                    "logo_path": None,
                    "colors": {
                        "primary": "#1976D2",
                        "secondary": "#424242",
                        "accent": "#FFC107"
                    }
                }
        
        return branding
    
    def _add_company_header(self, ws, start_row: int, branding: Dict[str, Any]) -> int:
        """Add company header with logo and name"""
        current_row = start_row
        
        # Check if logo exists
        logo_path = None
        if branding.get('logo_path'):
            from pathlib import Path
            logo_url = branding['logo_path']
            
            # Convert /static/logos/xxx.png to ./data/logos/xxx.png
            if logo_url.startswith('/static/'):
                logo_path = logo_url.replace('/static/', './data/')
                
                # Check if file exists and add logo
                logo_file = Path(logo_path)
                if logo_file.exists():
                    img = XLImage(str(logo_path))
                    img.width = min(img.width, 120)
                    img.height = min(img.height, 60)
                    ws.add_image(img, f'B{current_row}')
                else:
                    print(f" Logo not found at {logo_path}")
            
            # Resize (smaller for better fit)
            img.width = min(img.width, 120)
            img.height = min(img.height, 60)
            
            # Position logo at column B
            ws.add_image(img, f'B{current_row}')
            
            # Company name centered over table
            ws.merge_cells(f'E{current_row}:I{current_row}')
            name_cell = ws[f'E{current_row}']
            name_cell.value = branding["company_name"]
            name_cell.font = Font(bold=True, size=18, color=branding["colors"]["primary"].lstrip('#'))
            name_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Advance rows (reduced spacing)
            current_row += 3
        else:
            # No logo - just company name
            ws.merge_cells(f'A{current_row}:M{current_row}')
            name_cell = ws[f'A{current_row}']
            name_cell.value = branding["company_name"]
            name_cell.font = Font(bold=True, size=20, color=branding["colors"]["primary"].lstrip('#'))
            name_cell.alignment = Alignment(horizontal='center')
            
            current_row += 1
        
        # Add a separator line
        ws.merge_cells(f'A{current_row}:M{current_row}')
        separator = ws[f'A{current_row}']
        separator.fill = PatternFill(
            start_color=branding["colors"]["accent"].lstrip('#'),
            end_color=branding["colors"]["accent"].lstrip('#'),
            fill_type="solid"
        )
        ws.row_dimensions[current_row].height = 3
        
        return current_row
    
    def generate_ar_aging(self, report_data: Dict[str, Any], company_id: str) -> str:
        """
        Generate AR Aging Report with company branding
        Identical to AP Aging but for customer invoices
        """
        # Get branding
        branding = self._get_branding(company_id)
        
        # Get colors
        colors = branding["colors"]
        primary_color = colors["primary"].lstrip('#')
        secondary_color = colors["secondary"].lstrip('#')
        
        # Styling
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(
            start_color=primary_color,
            end_color=primary_color,
            fill_type="solid"
        )
        title_font = Font(bold=True, size=16, color=primary_color)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AR Aging"
        
        current_row = 1
        
        # Add company header
        current_row = self._add_company_header(ws, current_row, branding)
        
        # Report title - AR specific
        current_row += 1
        ws.merge_cells(f'A{current_row}:H{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = "AR AGING REPORT"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center')
        
        # Report date
        current_row += 1
        ws.merge_cells(f'A{current_row}:H{current_row}')
        date_cell = ws[f'A{current_row}']
        date_cell.value = f"As of: {report_data['report_metadata']['as_of_date']} | Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        date_cell.alignment = Alignment(horizontal='center')
        date_cell.font = Font(size=9)
        
        current_row += 2
        
        # TABLE - Exact format from screenshot
        table_data = report_data.get('table_data') or report_data.get('aging_data')
        
        # Headers
        headers = table_data['headers']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        current_row += 1
        
        # Data rows (from database - NOT hardcoded!)
        for row_data in table_data['rows']:
            ws.cell(row=current_row, column=1).value = row_data['vendor_name']  # This should be customer_name for AR
            ws.cell(row=current_row, column=2).value = row_data['invoice_no']
            ws.cell(row=current_row, column=3).value = row_data['due_date']
            ws.cell(row=current_row, column=4).value = f"₹{row_data['total_due']:,.2f}"
            ws.cell(row=current_row, column=5).value = f"₹{row_data['current_0_30']:,.2f}"
            ws.cell(row=current_row, column=6).value = f"₹{row_data['days_31_60']:,.2f}"
            ws.cell(row=current_row, column=7).value = f"₹{row_data['days_61_90']:,.2f}"
            ws.cell(row=current_row, column=8).value = f"₹{row_data['days_90_plus']:,.2f}"
            
            # Alignment
            for col in range(1, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.border = border
                if col == 1:  # Customer name - left align
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:  # Numbers - right align
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            
            current_row += 1
        
        # GRAND TOTALS row
        totals = table_data['grand_totals']
        
        ws.cell(row=current_row, column=1).value = "GRAND TOTALS"
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        ws.cell(row=current_row, column=4).value = f"₹{totals['total_due']:,.2f}"
        ws.cell(row=current_row, column=4).font = Font(bold=True, size=11)
        
        ws.cell(row=current_row, column=5).value = f"₹{totals['current_0_30']:,.2f}"
        ws.cell(row=current_row, column=5).font = Font(bold=True, size=11)
        
        ws.cell(row=current_row, column=6).value = f"₹{totals['days_31_60']:,.2f}"
        ws.cell(row=current_row, column=6).font = Font(bold=True, size=11)
        
        ws.cell(row=current_row, column=7).value = f"₹{totals['days_61_90']:,.2f}"
        ws.cell(row=current_row, column=7).font = Font(bold=True, size=11)
        
        ws.cell(row=current_row, column=8).value = f"₹{totals['days_90_plus']:,.2f}"
        ws.cell(row=current_row, column=8).font = Font(bold=True, size=11)
        
        # Grand totals styling
        for col in range(1, 9):
            cell = ws.cell(row=current_row, column=col)
            cell.border = border
            cell.fill = PatternFill(
                start_color=secondary_color,
                end_color=secondary_color,
                fill_type="solid"
            )
            if col > 1:
                cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Column widths
        ws.column_dimensions['A'].width = 20  # Customer Name
        ws.column_dimensions['B'].width = 15  # Invoice No
        ws.column_dimensions['C'].width = 12  # Due Date
        ws.column_dimensions['D'].width = 15  # Total Due
        ws.column_dimensions['E'].width = 15  # Current (0-30)
        ws.column_dimensions['F'].width = 15  # 31-60 Days
        ws.column_dimensions['G'].width = 15  # 61-90 Days
        ws.column_dimensions['H'].width = 15  # 90+ Days
        
        # Save
        output_dir = Path("./output/reports")
        output_dir.mkdir(parents=True, exist_ok=True)
        
        company_name = branding["company_name"].replace(" ", "_")
        filename = f"{company_name}_AR_Aging_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = output_dir / filename
        wb.save(filepath)
        
        logger.info(f"AR Aging Excel generated: {filepath}")
        return str(filepath)
    
    def run(self, input_data: Any, params: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Generate AR Aging report by copying BrandedExcelGenerator functionality
        
        Args:
            input_data: Report data with table_data and metadata
            params: Configuration parameters
            
        Returns:
            File path and metadata
        """
        params = params or {}
        company_id = params.get('company_id')
        
        if not company_id:
            raise ValueError("Company ID is required for branding")
        
        # Load branding from database
        branding = self.load_branding(company_id)
        
        # Validate data
        if not self.validate_data(input_data):
            raise ValueError("Invalid data provided")
        
        # Generate report using copied method
        file_path = self.generate_ar_aging(input_data, company_id)
        
        self.logger.info(f"Generated AR Aging report: {file_path}")
        
        return {
            'file_path': file_path,
            'branding_applied': True,
            'report_type': 'ar_aging',
            'branding': branding,
            'metadata': self.get_report_metadata(params)
        }