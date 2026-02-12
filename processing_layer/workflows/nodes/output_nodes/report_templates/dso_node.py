"""
DSO Report Node
Copies generate_dso_report method from BrandedExcelGenerator
"""

from pathlib import Path
from datetime import datetime
from typing import Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from processing_layer.workflows.nodes.output_nodes.base_template_node import BaseTemplateNode, register_node
from shared.branding.company_branding import CompanyBrandingManager
from shared.config.logging_config import get_logger


logger = get_logger(__name__)


@register_node
class DSOReportNode(BaseTemplateNode):
    """
    DSO Report Node
    Copies generate_dso_report method from BrandedExcelGenerator
    """
    
    name = "DSO Report"
    category = "output"
    description = "Generates DSO report by copying BrandedExcelGenerator functionality"
    
    input_schema = {
        "dso_data": {"type": "object", "required": True},
        "company_id": {"type": "string", "required": True}
    }
    
    output_schema = {
        "file_path": {"type": "string", "description": "Path to generated Excel file"},
        "branding_applied": {"type": "boolean", "description": "Whether branding was applied"},
        "report_type": {"type": "string", "description": "Type of report generated"}
    }
    
    ALLOWED_PARAMS = ["company_id"]
    
    def __init__(self):
        super().__init__()
    
    def _get_branding(self, company_id: str) -> Dict[str, Any]:
        """Get company branding configuration"""
        try:
            # Get database connection
            from sqlalchemy import create_engine, text
            from sqlalchemy.orm import sessionmaker
            import os
            
            db_url = f"postgresql://{os.getenv('DB_USER', 'postgres')}:{os.getenv('DB_PASSWORD', 'postgres')}@{os.getenv('DB_HOST', 'localhost')}:{os.getenv('DB_PORT', '5432')}/{os.getenv('DB_NAME', 'financial_automation')}"
            engine = create_engine(db_url)
            Session = sessionmaker(bind=engine)
            db_session = Session()
            
            # Get user's company
            query = text("""
                SELECT c.name, c.logo_url, c.primary_color, c.secondary_color
                FROM users u
                JOIN companies c ON u.company_id = c.id
                WHERE u.id = :user_id
            """)
            result = db_session.execute(query, {"user_id": company_id}).first()
            db_session.close()
            
            if result:
                return {
                    "company_name": result[0] or "Company",
                    "logo_path": result[1],
                    "colors": {
                        "primary": result[2] or "#1976D2",
                        "secondary": result[3] or "#424242",
                        "accent": "#FFC107"
                    }
                }
            else:
                raise Exception("No user found")
                
        except Exception:
            # Fallback to default branding
            return {
                "company_name": "Company",
                "logo_path": None,
                "colors": {
                    "primary": "#1976D2",
                    "secondary": "#424242",
                    "accent": "#FFC107"
                }
            }
    
    def _add_company_header(self, ws, start_row: int, branding: Dict[str, Any]) -> int:
        """Add company header with logo and name"""
        current_row = start_row
        
        # Check if logo exists
        if branding.get('logo_path'):
            from pathlib import Path
            logo_url = branding['logo_path']
            
            # Convert /static/logos/xxx.png to ./data/logos/xxx.png
            if logo_url.startswith('/static/'):
                logo_path = logo_url.replace('/static/', './data/')
                
                # Check if file exists and add logo
                logo_file = Path(logo_path)
                if logo_file.exists():
                    img = XLImage(str(logo_path))
                    img.width = min(img.width, 120)
                    img.height = min(img.height, 60)
                    ws.add_image(img, f'B{current_row}')
            
            # Company name centered over table
            ws.merge_cells(f'E{current_row}:I{current_row}')
            name_cell = ws[f'E{current_row}']
            name_cell.value = branding["company_name"]
            name_cell.font = Font(bold=True, size=18, color=branding["colors"]["primary"].lstrip('#'))
            name_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 3
        else:
            # No logo - just company name
            ws.merge_cells(f'A{current_row}:M{current_row}')
            name_cell = ws[f'A{current_row}']
            name_cell.value = branding["company_name"]
            name_cell.font = Font(bold=True, size=20, color=branding["colors"]["primary"].lstrip('#'))
            name_cell.alignment = Alignment(horizontal='center')
            
            current_row += 1
        
        # Add a separator line
        ws.merge_cells(f'A{current_row}:M{current_row}')
        separator = ws[f'A{current_row}']
        separator.fill = PatternFill(
            start_color=branding["colors"]["accent"].lstrip('#'),
            end_color=branding["colors"]["accent"].lstrip('#'),
            fill_type="solid"
        )
        ws.row_dimensions[current_row].height = 3
        
        return current_row
    
    def generate_dso_report(self, dso_data: Dict[str, Any], company_id: str) -> str:
        """
        Generate DSO Analysis Report with company branding
        """
        # Get branding
        branding = self._get_branding(company_id)
        
        # Get colors
        colors = branding["colors"]
        primary_color = colors["primary"].lstrip('#')
        secondary_color = colors["secondary"].lstrip('#')
        
        # Styling
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(
            start_color=primary_color,
            end_color=primary_color,
            fill_type="solid"
        )
        title_font = Font(bold=True, size=16, color=primary_color)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DSO Analysis"
        
        current_row = 1
        
        # Add company header
        current_row = self._add_company_header(ws, current_row, branding)
        
        # Report title
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = "DAYS SALES OUTSTANDING (DSO) ANALYSIS"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center')
        
        # Report metadata
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        meta_cell = ws[f'A{current_row}']
        meta_cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        meta_cell.alignment = Alignment(horizontal='center')
        meta_cell.font = Font(italic=True, size=10)
        
        # DSO Summary Section
        current_row += 2
        
        # Extract actual DSO data from the nested structure
        summary_data = dso_data
        if 'data' in dso_data:
            summary_data = dso_data['data']
        
        # Summary header
        ws.merge_cells(f'A{current_row}:F{current_row}')
        summary_header = ws[f'A{current_row}']
        summary_header.value = "DSO PERFORMANCE SUMMARY"
        summary_header.font = Font(bold=True, size=12, color=primary_color)
        summary_header.alignment = Alignment(horizontal='center')
        
        # Summary metrics
        current_row += 1
        
        # DSO Value with performance indicator
        ws[f'A{current_row}'] = f"DSO: {summary_data.get('dso', 0)} days"
        ws[f'A{current_row}'].font = Font(bold=True, size=14)
        
        # Performance indicator
        performance = summary_data.get('performance', 'Unknown')
        category = summary_data.get('category', 'unknown')
        
        ws[f'D{current_row}'] = f"Performance: {performance}"
        ws[f'D{current_row}'].font = Font(bold=True, size=12)
        
        # Color code the performance
        perf_cell = ws[f'D{current_row}']
        if category == 'success':
            perf_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        elif category == 'warning':
            perf_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        else:
            perf_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        current_row += 1
        
        # Key metrics
        metrics = [
            ("Total Sales", f"₹{summary_data.get('total_sales', 0):,.2f}"),
            ("Average AR", f"₹{summary_data.get('average_ar', 0):,.2f}"),
            ("Outstanding AR", f"₹{summary_data.get('outstanding_ar', 0):,.2f}"),
            ("Invoice Count", summary_data.get('invoice_count', 0)),
            ("Paid Invoices", summary_data.get('paid_invoices', 0)),
            ("Unpaid Invoices", summary_data.get('unpaid_invoices', 0))
        ]
        
        for i, (label, value) in enumerate(metrics):
            row_offset = current_row + i
            ws[f'A{row_offset}'] = label
            ws[f'A{row_offset}'].font = Font(bold=True)
            ws[f'B{row_offset}'] = value
            ws[f'B{row_offset}'].alignment = Alignment(horizontal='right')
        
        current_row += len(metrics) + 1
        
        # Collection Analysis Section
        ws.merge_cells(f'A{current_row}:F{current_row}')
        analysis_header = ws[f'A{current_row}']
        analysis_header.value = "COLLECTION ANALYSIS"
        analysis_header.font = Font(bold=True, size=12, color=primary_color)
        analysis_header.alignment = Alignment(horizontal='center')
        
        current_row += 1
        
        # Collection efficiency
        collection_efficiency = (summary_data.get('paid_invoices', 0) / summary_data.get('invoice_count', 1) * 100) if summary_data.get('invoice_count', 0) > 0 else 0
        
        ws[f'A{current_row}'] = "Collection Efficiency"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = f"{collection_efficiency:.1f}%"
        ws[f'B{current_row}'].alignment = Alignment(horizontal='right')
        
        current_row += 1
        
        # Average collection period
        ws[f'A{current_row}'] = "Average Collection Period"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = f"{summary_data.get('dso', 0):.1f} days"
        ws[f'B{current_row}'].alignment = Alignment(horizontal='right')
        
        current_row += 2
        
        # Recommendations Section
        ws.merge_cells(f'A{current_row}:F{current_row}')
        rec_header = ws[f'A{current_row}']
        rec_header.value = "COLLECTION RECOMMENDATIONS"
        rec_header.font = Font(bold=True, size=12, color=primary_color)
        rec_header.alignment = Alignment(horizontal='center')
        
        current_row += 1
        
        recommendations = self._get_dso_recommendations(summary_data)
        
        for i, rec in enumerate(recommendations):
            ws[f'A{current_row + i}'] = f"• {rec}"
            ws[f'A{current_row + i}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        current_row += len(recommendations) + 2
        
        # Invoice Details Section
        ws.merge_cells(f'A{current_row}:F{current_row}')
        details_header = ws[f'A{current_row}']
        details_header.value = "INVOICE DETAILS"
        details_header.font = Font(bold=True, size=12, color=primary_color)
        details_header.alignment = Alignment(horizontal='center')
        
        current_row += 2
        
        # Table headers
        headers = ["Invoice No.", "Customer", "Invoice Date", "Due Date", "Amount", "Outstanding", "Days Outstanding"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        current_row += 1
        
        # Note: Since DSO agent doesn't return individual invoice details,
        # we'll create a summary table or leave it for future enhancement
        ws.merge_cells(f'A{current_row}:G{current_row}')
        note_cell = ws[f'A{current_row}']
        note_cell.value = "Note: Detailed invoice breakdown available in AR Aging Report"
        note_cell.font = Font(italic=True, size=10)
        note_cell.alignment = Alignment(horizontal='center')
        
        # Column widths
        ws.column_dimensions['A'].width = 15  # Invoice No
        ws.column_dimensions['B'].width = 20  # Customer
        ws.column_dimensions['C'].width = 12  # Invoice Date
        ws.column_dimensions['D'].width = 12  # Due Date
        ws.column_dimensions['E'].width = 15  # Amount
        ws.column_dimensions['F'].width = 15  # Outstanding
        ws.column_dimensions['G'].width = 15  # Days Outstanding
        
        # Save
        output_dir = Path("./output/reports")
        output_dir.mkdir(parents=True, exist_ok=True)
        
        company_name = branding["company_name"].replace(" ", "_")
        filename = f"{company_name}_DSO_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = output_dir / filename
        wb.save(filepath)
        
        logger.info(f"DSO Analysis Excel generated: {filepath}")
        return str(filepath)
    
    def _get_dso_recommendations(self, dso_data: Dict[str, Any]) -> list:
        """
        Generate collection recommendations based on DSO performance
        """
        dso = dso_data.get('dso', 0)
        performance = dso_data.get('performance', 'Unknown')
        collection_efficiency = (dso_data.get('paid_invoices', 0) / dso_data.get('invoice_count', 1) * 100) if dso_data.get('invoice_count', 0) > 0 else 0
        
        recommendations = []
        
        if performance == "Excellent":
            recommendations.append("DSO performance is excellent. Continue current collection practices.")
            recommendations.append("Consider offering early payment discounts to maintain good relationships.")
            recommendations.append("Review credit terms for potential optimization opportunities.")
        
        elif performance == "Good":
            recommendations.append("DSO performance is good but can be improved.")
            recommendations.append("Implement stricter follow-up procedures for invoices over 30 days.")
            recommendations.append("Review customer credit limits and payment terms.")
        
        elif performance == "Fair":
            recommendations.append("DSO performance needs attention. Implement immediate improvements.")
            recommendations.append("Strengthen collection follow-up process for overdue accounts.")
            recommendations.append("Review and tighten credit approval policies.")
            recommendations.append("Consider offering payment plans for large outstanding balances.")
        
        else:  # Needs Improvement
            recommendations.append("DSO performance is poor. Immediate action required.")
            recommendations.append("Implement aggressive collection strategies for overdue accounts.")
            recommendations.append("Review and potentially restrict credit terms for high-risk customers.")
            recommendations.append("Consider engaging collection agencies for severely delinquent accounts.")
            recommendations.append("Analyze root causes of slow collections and address systemic issues.")
        
        # Additional recommendations based on collection efficiency
        if collection_efficiency < 70:
            recommendations.append("Collection efficiency is low. Focus on improving payment collection rates.")
        
        if dso_data.get('unpaid_invoices', 0) > 0:
            recommendations.append("Prioritize collection efforts on unpaid invoices to improve cash flow.")
        
        return recommendations
    
    def run(self, input_data: Any, params: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Generate DSO report by copying BrandedExcelGenerator functionality
        
        Args:
            input_data: DSO calculation result
            params: Configuration parameters
            
        Returns:
            File path and metadata
        """
        params = params or {}
        company_id = params.get('company_id')
        
        if not company_id:
            raise ValueError("Company ID is required for branding")
        
        # Load branding from database
        branding = self.load_branding(company_id)
        
        # Validate data
        if not self.validate_data(input_data):
            raise ValueError("Invalid data provided")
        
        # Generate report using copied method
        file_path = self.generate_dso_report(input_data, company_id)
        
        self.logger.info(f"Generated DSO report: {file_path}")
        
        return {
            'file_path': file_path,
            'branding_applied': True,
            'report_type': 'dso',
            'branding': branding,
            'metadata': self.get_report_metadata(params)
        }