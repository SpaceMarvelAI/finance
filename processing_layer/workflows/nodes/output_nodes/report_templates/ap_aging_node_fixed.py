"""
AP Aging Report Node
Copies generate_ap_aging method from BrandedExcelGenerator
"""

from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from processing_layer.workflows.nodes.output_nodes.base_template_node import BaseTemplateNode, register_node
from shared.branding.company_branding import CompanyBrandingManager
from shared.config.logging_config import get_logger


logger = get_logger(__name__)


@register_node
class APAgingReportNode(BaseTemplateNode):
    """
    AP Aging Report Node
    Copies generate_ap_aging method from BrandedExcelGenerator
    """
    
    name = "AP Aging Report"
    category = "output"
    description = "Generates AP Aging report by copying BrandedExcelGenerator functionality"
    
    input_schema = {
        "data": {"type": "any", "required": True},
        "company_id": {"type": "string", "required": True}
    }
    
    output_schema = {
        "file_path": {"type": "string", "description": "Path to generated Excel file"},
        "branding_applied": {"type": "boolean", "description": "Whether branding was applied"},
        "report_type": {"type": "string", "description": "Type of report generated"}
    }
    
    ALLOWED_PARAMS = ["company_id", "as_of_date"]
    
    def __init__(self):
        super().__init__()
    
    def _get_branding(self, company_id: str) -> Dict[str, Any]:
        """Get company branding configuration from database"""
        from sqlalchemy import create_engine, text
        from sqlalchemy.orm import sessionmaker
        import os
        
        try:
            # Get database connection
            db_url = f"postgresql://{os.getenv('DB_USER', 'postgres')}:{os.getenv('DB_PASSWORD', 'postgres')}@{os.getenv('DB_HOST', 'localhost')}:{os.getenv('DB_PORT', '5432')}/{os.getenv('DB_NAME', 'financial_automation')}"
            engine = create_engine(db_url)
            Session = sessionmaker(bind=engine)
            db_session = Session()
            
            # Get company branding (not user)
            query = text("""
                SELECT c.name, c.logo_url, c.primary_color, c.secondary_color, c.accent_color, c.currency
                FROM companies c
                WHERE c.id = :company_id
            """)
            result = db_session.execute(query, {"company_id": company_id}).first()
            
            if result:
                branding = {
                    "company_name": result[0] or "Company",
                    "logo_path": result[1],
                    "colors": {
                        "primary": result[2] or "#1976D2",
                        "secondary": result[3] or "#424242",
                        "accent": result[4] or "#FFC107"
                    },
                    "currency": result[5] or "INR"
                }
            else:
                raise Exception("No company found")
                
            db_session.close()
            
        except Exception as e:
            print(f"⚠️  Could not load branding from database: {e}")
            # Fallback to file-based branding
            branding_manager = CompanyBrandingManager()
            branding = branding_manager.get_branding(company_id)
            
            if not branding:
                # Use default branding if none configured
                print(f"⚠️  No custom branding for company {company_id}, using defaults")
                branding = {
                    "company_name": "Company",
                    "logo_path": None,
                    "colors": {
                        "primary": "#1976D2",
                        "secondary": "#424242",
                        "accent": "#FFC107"
                    },
                    "currency": "INR"
                }
        
        return branding
    
    def _add_company_header(self, ws, start_row: int, branding: Dict[str, Any]) -> int:
        """Add company header with logo and name"""
        current_row = start_row
        
        # Check if logo exists
        logo_path = None
        if branding.get('logo_path'):
            from pathlib import Path
            logo_url = branding['logo_path']
            
            # Convert /static/logos/xxx.png to ./data/logos/xxx.png
            if logo_url.startswith('/static/'):
                logo_path = logo_url.replace('/static/', './data/')
                
                # Check if file exists and add logo
                logo_file = Path(logo_path)
                if logo_file.exists():
                    img = XLImage(str(logo_path))
                    # Resize (smaller for better fit)
                    img.width = min(img.width, 120)
                    img.height = min(img.height, 60)
                    
                    # Position logo at column B
                    ws.add_image(img, f'B{current_row}')
                else:
                    print(f" Logo not found at {logo_path}")
            
            # Company name centered over table
            ws.merge_cells(f'E{current_row}:I{current_row}')
            name_cell = ws[f'E{current_row}']
            name_cell.value = branding["company_name"]
            name_cell.font = Font(bold=True, size=18, color=branding["colors"]["primary"].lstrip('#'))
            name_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Advance rows (reduced spacing)
            current_row += 3
        else:
            # No logo - just company name
            ws.merge_cells(f'A{current_row}:M{current_row}')
            name_cell = ws[f'A{current_row}']
            name_cell.value = branding["company_name"]
            name_cell.font = Font(bold=True, size=20, color=branding["colors"]["primary"].lstrip('#'))
            name_cell.alignment = Alignment(horizontal='center')
            
            current_row += 1
        
        # Add a separator line
        ws.merge_cells(f'A{current_row}:M{current_row}')
        separator = ws[f'A{current_row}']
        separator.fill = PatternFill(
            start_color=branding["colors"]["accent"].lstrip('#'),
            end_color=branding["colors"]["accent"].lstrip('#'),
            fill_type="solid"
        )
        ws.row_dimensions[current_row].height = 3
        
        return current_row
    
    def generate_ap_aging(self, report_data: Dict[str, Any], company_id: str) -> str:
        """
        Generate AP Aging Report with company branding
        
        Table Format (EXACT):
        Vendor Name | Invoice No. | Due Date | Total Due | Current (0-30) | 31-60 Days | 61-90 Days | 90+ Days
        """
        # Get branding
        branding = self._get_branding(company_id)
        
        # Get colors
        colors = branding["colors"]
        primary_color = colors["primary"].lstrip('#')
        secondary_color = colors["secondary"].lstrip('#')
        
        # Styling
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(
            start_color=primary_color,
            end_color=primary_color,
            fill_type="solid"
        )
        title_font = Font(bold=True, size=16, color=primary_color)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AP Aging"
        
        current_row = 1
        
        # Add company header
        current_row = self._add_company_header(ws, current_row, branding)
        
        # Report title - support both AP and AR
        current_row += 1
        ws.merge_cells(f'A{current_row}:H{current_row}')
        title_cell = ws[f'A{current_row}']
        report_type = report_data.get('report_type', 'AP_AGING')
        if 'AR' in report_type.upper():
            title_cell.value = "AR AGING REPORT"
            ws.title = "AR Aging"
        else:
            title_cell.value = "AP AGING REPORT"
            ws.title = "AP Aging"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center')
        
        # Report date
        current_row += 1
        ws.merge_cells(f'A{current_row}:H{current_row}')
        date_cell = ws[f'A{current_row}']
        date_cell.value = f"As of: {report_data['report_metadata']['as_of_date']} | Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        date_cell.alignment = Alignment(horizontal='center')
        date_cell.font = Font(size=9)
        
        current_row += 2
        
        # TABLE - Exact format from screenshot
        table_data = report_data.get('table_data') or report_data.get('aging_data')
        
        # Headers
        headers = table_data['headers']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        current_row += 1
        
        # Data rows (from database - NOT hardcoded!)
        for row_data in table_data['rows']:
            ws.cell(row=current_row, column=1).value = row_data['vendor_name']
            ws.cell(row=current_row, column=2).value = row_data['invoice_no']
            ws.cell(row=current_row, column=3).value = row_data['due_date']
            ws.cell(row=current_row, column=4).value = f"₹{row_data['total_due']:,.2f}"
            ws.cell(row=current_row, column=5).value = f"₹{row_data['current_0_30']:,.2f}"
            ws.cell(row=current_row, column=6).value = f"₹{row_data['days_31_60']:,.2f}"
            ws.cell(row=current_row, column=7).value = f"₹{row_data['days_61_90']:,.2f}"
            ws.cell(row=current_row, column=8).value = f"₹{row_data['days_90_plus']:,.2f}"
            
            # Alignment
            for col in range(1, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.border = border
                if col == 1:  # Vendor name - left align
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:  # Numbers - right align
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            
            current_row += 1
        
        # GRAND TOTALS row
        totals = table_data['grand_totals']
        
        ws.cell(row=current_row, column=1).value = "GRAND TOTALS"
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        ws.cell(row=current_row, column=4).value = f"₹{totals['total_due']:,.2f}"
        ws.cell(row=current_row, column=4).font = Font(bold=True, size=11)
        
        ws.cell(row=current_row, column=5).value = f"₹{totals['current_0_30']:,.2f}"
        ws.cell(row=current_row, column=5).font = Font(bold=True, size=11)
        
        ws.cell(row=current_row, column=6).value = f"₹{totals['days_31_60']:,.2f}"
        ws.cell(row=current_row, column=6).font = Font(bold=True, size=11)
        
        ws.cell(row=current_row, column=7).value = f"₹{totals['days_61_90']:,.2f}"
        ws.cell(row=current_row, column=7).font = Font(bold=True, size=11)
        
        ws.cell(row=current_row, column=8).value = f"₹{totals['days_90_plus']:,.2f}"
        ws.cell(row=current_row, column=8).font = Font(bold=True, size=11)
        
        # Grand totals styling
        for col in range(1, 9):
            cell = ws.cell(row=current_row, column=col)
            cell.border = border
            cell.fill = PatternFill(
                start_color=secondary_color,
                end_color=secondary_color,
                fill_type="solid"
            )
            if col > 1:
                cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Column widths
        ws.column_dimensions['A'].width = 20  # Vendor Name
        ws.column_dimensions['B'].width = 15  # Invoice No
        ws.column_dimensions['C'].width = 12  # Due Date
        ws.column_dimensions['D'].width = 15  # Total Due
        ws.column_dimensions['E'].width = 15  # Current (0-30)
        ws.column_dimensions['F'].width = 15  # 31-60 Days
        ws.column_dimensions['G'].width = 15  # 61-90 Days
        ws.column_dimensions['H'].width = 15  # 90+ Days
        
        # Save
        output_dir = Path("./output/reports")
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Get company name from branding
        company_name = branding.get("company_name", "Company")
        if not company_name or company_name == "Company":
            # Fallback to a default name if not found
            company_name = "Company"
        
        # Clean company name for filename
        clean_company_name = company_name.replace(" ", "_").replace("/", "_").replace("\\", "_")
        
        report_type_str = report_data.get('report_type', 'AP_AGING')
        if 'AR' in report_type_str.upper():
            report_name = "AR_Aging"
            log_msg = "AR Aging Excel generated"
        else:
            report_name = "AP_Aging"
            log_msg = "AP Aging Excel generated"
        
        filename = f"{clean_company_name}_{report_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = output_dir / filename
        wb.save(filepath)
        
        logger.info(f"{log_msg}: {filepath}")
        return str(filepath)
    
    def run(self, input_data: Any, params: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Generate AP Aging report by copying BrandedExcelGenerator functionality
        
        Args:
            input_data: Raw invoice data or report data
            params: Configuration parameters
            
        Returns:
            File path and metadata
        """
        params = params or {}
        company_id = params.get('company_id')
        
        if not company_id:
            raise ValueError("Company ID is required for branding")
        
        # Load branding from database
        branding = self._get_branding(company_id)
        
        # Check if input_data is raw invoice data (list of dicts) or already formatted report data
        if isinstance(input_data, list) and len(input_data) > 0 and isinstance(input_data[0], dict):
            # Raw invoice data - convert to report format
            report_data = self._convert_invoices_to_report_data(input_data, company_id)
        elif isinstance(input_data, dict) and 'data' in input_data:
            # Already formatted report data
            report_data = input_data['data']
        else:
            # Handle other cases
            report_data = input_data
        
        # Generate report using copied method
        file_path = self.generate_ap_aging(report_data, company_id)
        
        self.logger.info(f"Generated AP Aging report: {file_path}")
        
        return {
            'file_path': file_path,
            'branding_applied': True,
            'report_type': 'ap_aging',
            'branding': branding,
            'metadata': self.get_report_metadata(params)
        }
    
    def _convert_invoices_to_report_data(self, invoices: list, company_id: str) -> dict:
        """Convert raw invoice data to report format"""
        from datetime import datetime, timedelta
        
        today = datetime.now().date()
        table_rows = []
        
        for invoice in invoices:
            # Calculate aging buckets
            due_date = self._parse_date(invoice.get('due_date'))
            if due_date:
                days_overdue = (today - due_date).days
            else:
                days_overdue = 0
            
            total_due = float(invoice.get('outstanding_amount', invoice.get('total_amount', 0)))
            
            # Assign to aging buckets
            if days_overdue <= 30:
                current_0_30 = total_due
                days_31_60 = 0
                days_61_90 = 0
                days_90_plus = 0
            elif days_overdue <= 60:
                current_0_30 = 0
                days_31_60 = total_due
                days_61_90 = 0
                days_90_plus = 0
            elif days_overdue <= 90:
                current_0_30 = 0
                days_31_60 = 0
                days_61_90 = total_due
                days_90_plus = 0
            else:
                current_0_30 = 0
                days_31_60 = 0
                days_61_90 = 0
                days_90_plus = total_due
            
            table_rows.append({
                'vendor_name': invoice.get('vendor_name', 'Unknown'),
                'invoice_no': invoice.get('invoice_number', ''),
                'due_date': invoice.get('due_date', ''),
                'total_due': total_due,
                'current_0_30': current_0_30,
                'days_31_60': days_31_60,
                'days_61_90': days_61_90,
                'days_90_plus': days_90_plus
            })
        
        # Calculate totals
        grand_totals = {
            'total_due': sum(row['total_due'] for row in table_rows),
            'current_0_30': sum(row['current_0_30'] for row in table_rows),
            'days_31_60': sum(row['days_31_60'] for row in table_rows),
            'days_61_90': sum(row['days_61_90'] for row in table_rows),
            'days_90_plus': sum(row['days_90_plus'] for row in table_rows)
        }
        
        return {
            'report_type': 'AP_AGING',
            'report_metadata': {
                'as_of_date': today.strftime('%Y-%m-%d')
            },
            'table_data': {
                'headers': ["Vendor Name", "Invoice No.", "Due Date", "Total Due", "Current (0-30)", "31-60 Days", "61-90 Days", "90+ Days"],
                'rows': table_rows,
                'grand_totals': grand_totals
            }
        }
    
    def _parse_date(self, date_str: str) -> Optional[datetime.date]:
        """Parse date string to date object"""
        if not date_str:
            return None
        
        try:
            from datetime import datetime
            # Try common date formats
            for fmt in ['%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y', '%m/%d/%Y']:
                try:
                    return datetime.strptime(date_str, fmt).date()
                except ValueError:
                    continue
            return None
        except Exception:
            return None
