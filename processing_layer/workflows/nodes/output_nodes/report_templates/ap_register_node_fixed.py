"""
AP Register Report Node
Copies generate_ap_invoice_register method from BrandedExcelGenerator
"""

from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from processing_layer.workflows.nodes.output_nodes.base_template_node import BaseTemplateNode, register_node
from shared.branding.company_branding import CompanyBrandingManager
from shared.config.logging_config import get_logger


logger = get_logger(__name__)


@register_node
class APRegisterReportNode(BaseTemplateNode):
    """
    AP Register Report Node
    Copies generate_ap_invoice_register method from BrandedExcelGenerator
    """
    
    name = "AP Register Report"
    category = "output"
    description = "Generates AP Register report by copying BrandedExcelGenerator functionality"
    
    input_schema = {
        "report_data": {"type": "object", "required": True},
        "company_id": {"type": "string", "required": True}
    }
    
    output_schema = {
        "file_path": {"type": "string", "description": "Path to generated Excel file"},
        "branding_applied": {"type": "boolean", "description": "Whether branding was applied"},
        "report_type": {"type": "string", "description": "Type of report generated"}
    }
    
    ALLOWED_PARAMS = ["company_id"]
    
    def __init__(self):
        super().__init__()
    
    def _get_branding(self, company_id: str) -> Dict[str, Any]:
        """Get company branding configuration"""
        # Try to load branding from database first
        from sqlalchemy import create_engine, text
        from sqlalchemy.orm import sessionmaker
        import os
        
        try:
            # Get database connection
            db_url = f"postgresql://{os.getenv('DB_USER', 'postgres')}:{os.getenv('DB_PASSWORD', 'postgres')}@{os.getenv('DB_HOST', 'localhost')}:{os.getenv('DB_PORT', '5432')}/{os.getenv('DB_NAME', 'financial_automation')}"
            engine = create_engine(db_url)
            Session = sessionmaker(bind=engine)
            db_session = Session()
            
            # Get user's company
            query = text("""
                SELECT c.name, c.logo_url, c.primary_color, c.secondary_color
                FROM users u
                JOIN companies c ON u.company_id = c.id
                WHERE u.id = :user_id
            """)
            result = db_session.execute(query, {"user_id": company_id}).first()
            
            if result:
                branding = {
                    "company_name": result[0] or "Company",
                    "logo_path": result[1],
                    "colors": {
                        "primary": result[2] or "#1976D2",
                        "secondary": result[3] or "#424242",
                        "accent": "#FFC107"  # Default accent color
                    }
                }
            else:
                raise Exception("No user found")
                
            db_session.close()
            
        except Exception as e:
            print(f"⚠️  Could not load branding from database: {e}")
            # Fallback to file-based branding
            branding_manager = CompanyBrandingManager()
            branding = branding_manager.get_branding(company_id)
            
            if not branding:
                # Use default branding if none configured
                print(f"⚠️  No custom branding for user {company_id}, using defaults")
                branding = {
                    "company_name": "Company",
                    "logo_path": None,
                    "colors": {
                        "primary": "#1976D2",
                        "secondary": "#424242",
                        "accent": "#FFC107"
                    }
                }
        
        return branding
    
    def _add_company_header(self, ws, start_row: int, branding: Dict[str, Any]) -> int:
        """Add company header with logo and name"""
        current_row = start_row
        
        # Check if logo exists
        logo_path = None
        if branding.get('logo_path'):
            from pathlib import Path
            logo_url = branding['logo_path']
            
            # Convert /static/logos/xxx.png to ./data/logos/xxx.png
            if logo_url.startswith('/static/'):
                logo_path = logo_url.replace('/static/', './data/')
                
                # Check if file exists and add logo
                logo_file = Path(logo_path)
                if logo_file.exists():
                    img = XLImage(str(logo_path))
                    img.width = min(img.width, 120)
                    img.height = min(img.height, 60)
                    ws.add_image(img, f'B{current_row}')
                else:
                    print(f" Logo not found at {logo_path}")
            
            # Resize (smaller for better fit)
            img.width = min(img.width, 120)
            img.height = min(img.height, 60)
            
            # Position logo at column B
            ws.add_image(img, f'B{current_row}')
            
            # Company name centered over table
            ws.merge_cells(f'E{current_row}:I{current_row}')
            name_cell = ws[f'E{current_row}']
            name_cell.value = branding["company_name"]
            name_cell.font = Font(bold=True, size=18, color=branding["colors"]["primary"].lstrip('#'))
            name_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Advance rows (reduced spacing)
            current_row += 3
        else:
            # No logo - just company name
            ws.merge_cells(f'A{current_row}:M{current_row}')
            name_cell = ws[f'A{current_row}']
            name_cell.value = branding["company_name"]
            name_cell.font = Font(bold=True, size=20, color=branding["colors"]["primary"].lstrip('#'))
            name_cell.alignment = Alignment(horizontal='center')
            
            current_row += 1
        
        # Add a separator line
        ws.merge_cells(f'A{current_row}:M{current_row}')
        separator = ws[f'A{current_row}']
        separator.fill = PatternFill(
            start_color=branding["colors"]["accent"].lstrip('#'),
            end_color=branding["colors"]["accent"].lstrip('#'),
            fill_type="solid"
        )
        ws.row_dimensions[current_row].height = 3
        
        return current_row
    
    def generate_ap_invoice_register(self, report_data: Dict[str, Any], company_id: str) -> str:
        """
        Generate AP Invoice Register with company branding
        
        Args:
            report_data: Structured report data from database
            
        Returns:
            Path to generated Excel file
        """
        # Get branding
        branding = self._get_branding(company_id)
        
        # Get colors
        colors = branding["colors"]
        primary_color = colors["primary"].lstrip('#')
        secondary_color = colors["secondary"].lstrip('#')
        
        # Styling
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(
            start_color=primary_color,
            end_color=primary_color,
            fill_type="solid"
        )
        title_font = Font(bold=True, size=16, color=primary_color)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AP Invoice Register"
        
        current_row = 1
        
        # Add company header
        current_row = self._add_company_header(ws, current_row, branding)
        
        # Report title
        current_row += 1
        ws.merge_cells(f'A{current_row}:M{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = "AP INVOICE REGISTER REPORT"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center')
        
        # Report metadata
        current_row += 1
        ws.merge_cells(f'A{current_row}:M{current_row}')
        meta_cell = ws[f'A{current_row}']
        meta_cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        meta_cell.alignment = Alignment(horizontal='center')
        meta_cell.font = Font(italic=True, size=10)
        
        # Summary section
        current_row += 2
        summary = report_data['summary']
        
        summary_fill = PatternFill(
            start_color=secondary_color,
            end_color=secondary_color,
            fill_type="solid"
        )
        
        ws[f'A{current_row}'] = f"Total Invoices: {summary['total_invoices']}"
        ws[f'A{current_row}'].fill = summary_fill
        ws[f'A{current_row}'].font = Font(bold=True)
        
        ws[f'D{current_row}'] = f"Paid: {summary['paid_count']}"
        ws[f'D{current_row}'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        ws[f'G{current_row}'] = f"Partial: {summary['partial_count']}"
        ws[f'G{current_row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        ws[f'J{current_row}'] = f"Unpaid: {summary['unpaid_count']}"
        ws[f'J{current_row}'].fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # Table headers
        current_row += 2
        headers = [
            "Trans ID", "Vendor Name", "Vendor ID", "Invoice No.",
            "Invoice Date", "Due Date", "Description",
            "Net Amt", "Tax Amt", "Sub Total", "Paid Amt",
            "Outstanding", "Status"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        current_row += 1
        for invoice in report_data['invoices']:
            ws.cell(row=current_row, column=1).value = invoice['trans_id']
            ws.cell(row=current_row, column=2).value = invoice['vendor_name']
            ws.cell(row=current_row, column=3).value = invoice['vendor_id']
            ws.cell(row=current_row, column=4).value = invoice['invoice_no']
            ws.cell(row=current_row, column=5).value = invoice['invoice_date']
            ws.cell(row=current_row, column=6).value = invoice['due_date']
            ws.cell(row=current_row, column=7).value = invoice['description']
            ws.cell(row=current_row, column=8).value = f"₹{invoice['net_amt']:,.2f}"
            ws.cell(row=current_row, column=9).value = f"₹{invoice['tax_amt']:,.2f}"
            ws.cell(row=current_row, column=10).value = f"₹{invoice['sub_total']:,.2f}"
            ws.cell(row=current_row, column=11).value = f"₹{invoice['paid_amt']:,.2f}"
            ws.cell(row=current_row, column=12).value = f"₹{invoice['outstanding']:,.2f}"
            ws.cell(row=current_row, column=13).value = invoice['status']
            
            # Status color coding
            status_cell = ws.cell(row=current_row, column=13)
            if invoice['status'] == 'Paid':
                status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif invoice['status'] == 'Partial':
                status_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            # Borders
            for col in range(1, 14):
                ws.cell(row=current_row, column=col).border = border
                ws.cell(row=current_row, column=col).alignment = Alignment(vertical='center')
            
            current_row += 1
        
        # Totals row
        totals = report_data.get('totals') or {}
        if not totals or not totals.get('invoice_amt'):
            summary = report_data.get('summary', {})
            totals = {
                'invoice_amt': summary.get('total_amount', 0),
                'tax_amt': 0,  # Will calculate from invoices
                'net_amt': summary.get('total_amount', 0),
                'paid_amt': summary.get('total_paid', 0),
                'outstanding': summary.get('total_outstanding', 0)
            }
        ws.cell(row=current_row, column=1).value = "TOTALS"
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        
        ws.cell(row=current_row, column=8).value = f"₹{totals['invoice_amt']:,.2f}"
        ws.cell(row=current_row, column=8).value = f"₹{totals['invoice_amt']:,.2f}"
        ws.cell(row=current_row, column=8).font = Font(bold=True)
        ws.cell(row=current_row, column=9).value = f"₹{totals['tax_amt']:,.2f}"
        ws.cell(row=current_row, column=9).font = Font(bold=True)
        ws.cell(row=current_row, column=10).value = f"₹{totals['net_amt']:,.2f}"
        ws.cell(row=current_row, column=10).font = Font(bold=True)
        ws.cell(row=current_row, column=11).value = f"₹{totals['paid_amt']:,.2f}"
        ws.cell(row=current_row, column=11).font = Font(bold=True)
        ws.cell(row=current_row, column=12).value = f"₹{totals['outstanding']:,.2f}"
        ws.cell(row=current_row, column=12).font = Font(bold=True)
        
        # Borders on totals
        for col in range(1, 14):
            ws.cell(row=current_row, column=col).border = border
            ws.cell(row=current_row, column=col).fill = PatternFill(
                start_color=secondary_color,
                end_color=secondary_color,
                fill_type="solid"
            )
        
        # Column widths
        column_widths = [10, 20, 12, 15, 12, 12, 25, 14, 12, 12, 12, 14, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Save
        output_dir = Path("./output/reports")
        output_dir.mkdir(parents=True, exist_ok=True)
        
        company_name = branding["company_name"].replace(" ", "_")
        filename = f"{company_name}_AP_Invoice_Register_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = output_dir / filename
        wb.save(filepath)
        
        logger.info(f"Branded Excel generated: {filepath}")
        return str(filepath)
    
    def run(self, input_data: Any, params: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Generate AP Register report by copying BrandedExcelGenerator functionality
        
        Args:
            input_data: Report data with invoices and summary
            params: Configuration parameters
            
        Returns:
            File path and metadata
        """
        params = params or {}
        company_id = params.get('company_id')
        
        if not company_id:
            raise ValueError("Company ID is required for branding")
        
        # Load branding from database
        branding = self.load_branding(company_id)
        
        # Validate data
        if not self.validate_data(input_data):
            raise ValueError("Invalid data provided")
        
        # Extract the actual report data from input_data
        # input_data should be {'data': report_data, 'company_id': company_id}
        report_data = input_data.get('data')
        if not report_data:
            raise ValueError("Missing report data")
        
        # Generate report using copied method
        file_path = self.generate_ap_invoice_register(report_data, company_id)
        
        self.logger.info(f"Generated AP Register report: {file_path}")
        
        return {
            'file_path': file_path,
            'branding_applied': True,
            'report_type': 'ap_register',
            'branding': branding,
            'metadata': self.get_report_metadata(params)
        }
