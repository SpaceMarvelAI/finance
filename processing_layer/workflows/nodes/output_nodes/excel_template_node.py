"""
Excel Template Node
Generates branded Excel reports using openpyxl
"""

from typing import Dict, Any, List, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
from processing_layer.workflows.nodes.output_nodes.base_template_node import BaseTemplateNode, register_node
from data_layer.database.database_manager import get_database


@register_node
class ExcelTemplateNode(BaseTemplateNode):
    """
    Excel report generator with company branding
    Supports multiple report types with dynamic templates
    """
    
    name = "Excel Template"
    category = "output"
    description = "Generates branded Excel reports with company branding"
    
    input_schema = {
        "data": {"type": "any", "required": True},
        "company_id": {"type": "string", "required": True},
        "report_type": {"type": "string", "enum": ["ap_aging", "ar_aging", "ap_register", "ar_register", "ap_overdue", "dso"]}
    }
    
    output_schema = {
        "file_path": {"type": "string", "description": "Path to generated Excel file"},
        "branding_applied": {"type": "boolean", "description": "Whether branding was applied"},
        "report_type": {"type": "string", "description": "Type of report generated"}
    }
    
    ALLOWED_PARAMS = ["company_id", "report_type", "include_charts", "template_style", "as_of_date"]
    
    def __init__(self):
        super().__init__()
        self.db = get_database()
    
    def apply_branding(self, workbook, branding: Dict[str, Any]):
        """
        Apply company branding to Excel workbook
        
        Args:
            workbook: openpyxl workbook object
            branding: Branding configuration from database
        """
        # Set default colors if not provided
        colors = branding.get('colors', {})
        primary_color = colors.get('primary', '1976D2')  # Default blue
        secondary_color = colors.get('secondary', '424242')  # Default gray
        accent_color = colors.get('accent', 'FFC107')  # Default amber
        
        # Clean color codes (remove # if present)
        primary_color = primary_color.replace('#', '')
        secondary_color = secondary_color.replace('#', '')
        accent_color = accent_color.replace('#', '')
        
        # Apply branding to all sheets
        for sheet in workbook.worksheets:
            self._apply_sheet_branding(sheet, primary_color, secondary_color, accent_color, branding)
    
    def _apply_sheet_branding(self, sheet, primary_color: str, secondary_color: str, accent_color: str, branding: Dict[str, Any]):
        """Apply branding to individual sheet"""
        
        # Add company logo if available
        if branding.get('logo_path'):
            try:
                logo_path = branding['logo_path'].replace('/static/', './data/')
                if logo_path.startswith('./data/logos/'):
                    img = Image(logo_path)
                    img.width = 150
                    img.height = 50
                    sheet.add_image(img, 'A1')
            except Exception as e:
                self.logger.warning(f"Could not add logo: {e}")
        
        # Set column widths
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 15
        sheet.column_dimensions['G'].width = 15
        sheet.column_dimensions['H'].width = 15
        
        # Apply header styling
        header_row = 1 if branding.get('logo_path') else 1
        for cell in sheet[header_row]:
            if cell.value:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
    
    def add_company_header(self, workbook, branding: Dict[str, Any], report_title: str):
        """Add company header to Excel workbook"""
        sheet = workbook.active
        
        # Add company name and report title
        company_name = branding.get('company_name', 'Financial Automation System')
        currency = branding.get('currency', 'INR')
        
        # Position header based on logo
        start_row = 1
        if branding.get('logo_path'):
            start_row = 2
        
        # Company name
        sheet.cell(row=start_row, column=1, value=company_name).font = Font(size=16, bold=True)
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
        
        # Report title
        sheet.cell(row=start_row + 1, column=1, value=report_title).font = Font(size=14, bold=True)
        sheet.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=4)
        
        # Generated date
        sheet.cell(row=start_row + 2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        sheet.merge_cells(start_row=start_row + 2, start_column=1, end_row=start_row + 2, end_column=4)
        
        # Currency
        sheet.cell(row=start_row + 3, column=1, value=f"Currency: {currency}")
        sheet.merge_cells(start_row=start_row + 3, start_column=1, end_row=start_row + 3, end_column=4)
    
    def add_footer(self, workbook, branding: Dict[str, Any]):
        """Add footer to Excel workbook"""
        sheet = workbook.active
        last_row = sheet.max_row + 2
        
        # Footer text
        footer_text = f"Confidential - {branding.get('company_name', 'Financial Automation System')} - Generated by Financial Automation System"
        sheet.cell(row=last_row, column=1, value=footer_text).font = Font(size=10, italic=True)
        sheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=8)
    
    def render_ap_aging_template(self, data: Any, branding: Dict[str, Any]) -> openpyxl.Workbook:
        """Render AP Aging report template"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "AP Aging Report"
        
        # Add header
        self.add_company_header(workbook, branding, "AP Aging Report")
        
        # Add data table
        start_row = 5
        headers = ["Vendor Name", "Invoice No.", "Due Date", "Total Due", "Current (0-30)", "31-60 Days", "61-90 Days", "90+ Days"]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            sheet.cell(row=start_row, column=col, value=header)
        
        # Write data
        current_row = start_row + 1
        total_due = 0
        current_total = 0
        days_31_60_total = 0
        days_61_90_total = 0
        days_90_plus_total = 0
        
        # Handle both list of invoices and dict with invoices
        if isinstance(data, list):
            invoices = data
        elif isinstance(data, dict) and 'invoices' in data:
            invoices = data['invoices']
        else:
            invoices = []
        
        for invoice in invoices:
            vendor_name = invoice.get('vendor_name', 'Unknown')
            invoice_no = invoice.get('invoice_no', '')
            due_date = self.format_date(invoice.get('due_date'))
            total_due_amount = float(invoice.get('total_due', 0))
            
            # Get aging buckets
            current = float(invoice.get('current_0_30', 0))
            days_31_60 = float(invoice.get('days_31_60', 0))
            days_61_90 = float(invoice.get('days_61_90', 0))
            days_90_plus = float(invoice.get('days_90_plus', 0))
            
            # Write row
            sheet.cell(row=current_row, column=1, value=vendor_name)
            sheet.cell(row=current_row, column=2, value=invoice_no)
            sheet.cell(row=current_row, column=3, value=due_date)
            sheet.cell(row=current_row, column=4, value=total_due_amount)
            sheet.cell(row=current_row, column=5, value=current)
            sheet.cell(row=current_row, column=6, value=days_31_60)
            sheet.cell(row=current_row, column=7, value=days_61_90)
            sheet.cell(row=current_row, column=8, value=days_90_plus)
            
            # Update totals
            total_due += total_due_amount
            current_total += current
            days_31_60_total += days_31_60
            days_61_90_total += days_61_90
            days_90_plus_total += days_90_plus
            
            current_row += 1
        
        # Add totals row
        sheet.cell(row=current_row, column=1, value="TOTAL").font = Font(bold=True)
        sheet.cell(row=current_row, column=4, value=total_due).font = Font(bold=True)
        sheet.cell(row=current_row, column=5, value=current_total).font = Font(bold=True)
        sheet.cell(row=current_row, column=6, value=days_31_60_total).font = Font(bold=True)
        sheet.cell(row=current_row, column=7, value=days_61_90_total).font = Font(bold=True)
        sheet.cell(row=current_row, column=8, value=days_90_plus_total).font = Font(bold=True)
        
        # Apply branding
        self.apply_branding(workbook, branding)
        
        # Add footer
        self.add_footer(workbook, branding)
        
        return workbook
    
    def render_ar_aging_template(self, data: Any, branding: Dict[str, Any]) -> openpyxl.Workbook:
        """Render AR Aging report template"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "AR Aging Report"
        
        # Add header
        self.add_company_header(workbook, branding, "AR Aging Report")
        
        # Add data table
        start_row = 5
        headers = ["Customer Name", "Invoice No.", "Due Date", "Total Due", "Current (0-30)", "31-60 Days", "61-90 Days", "90+ Days"]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            sheet.cell(row=start_row, column=col, value=header)
        
        # Write data
        current_row = start_row + 1
        total_due = 0
        current_total = 0
        days_31_60_total = 0
        days_61_90_total = 0
        days_90_plus_total = 0
        
        # Handle both list of invoices and dict with invoices
        if isinstance(data, list):
            invoices = data
        elif isinstance(data, dict) and 'invoices' in data:
            invoices = data['invoices']
        else:
            invoices = []
        
        for invoice in invoices:
            customer_name = invoice.get('customer_name', 'Unknown')
            invoice_no = invoice.get('invoice_no', '')
            due_date = self.format_date(invoice.get('due_date'))
            total_due_amount = float(invoice.get('total_due', 0))
            
            # Get aging buckets
            current = float(invoice.get('current_0_30', 0))
            days_31_60 = float(invoice.get('days_31_60', 0))
            days_61_90 = float(invoice.get('days_61_90', 0))
            days_90_plus = float(invoice.get('days_90_plus', 0))
            
            # Write row
            sheet.cell(row=current_row, column=1, value=customer_name)
            sheet.cell(row=current_row, column=2, value=invoice_no)
            sheet.cell(row=current_row, column=3, value=due_date)
            sheet.cell(row=current_row, column=4, value=total_due_amount)
            sheet.cell(row=current_row, column=5, value=current)
            sheet.cell(row=current_row, column=6, value=days_31_60)
            sheet.cell(row=current_row, column=7, value=days_61_90)
            sheet.cell(row=current_row, column=8, value=days_90_plus)
            
            # Update totals
            total_due += total_due_amount
            current_total += current
            days_31_60_total += days_31_60
            days_61_90_total += days_61_90
            days_90_plus_total += days_90_plus
            
            current_row += 1
        
        # Add totals row
        sheet.cell(row=current_row, column=1, value="TOTAL").font = Font(bold=True)
        sheet.cell(row=current_row, column=4, value=total_due).font = Font(bold=True)
        sheet.cell(row=current_row, column=5, value=current_total).font = Font(bold=True)
        sheet.cell(row=current_row, column=6, value=days_31_60_total).font = Font(bold=True)
        sheet.cell(row=current_row, column=7, value=days_61_90_total).font = Font(bold=True)
        sheet.cell(row=current_row, column=8, value=days_90_plus_total).font = Font(bold=True)
        
        # Apply branding
        self.apply_branding(workbook, branding)
        
        # Add footer
        self.add_footer(workbook, branding)
        
        return workbook
    
    def render_ap_register_template(self, data: Any, branding: Dict[str, Any]) -> openpyxl.Workbook:
        """Render AP Register report template"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "AP Register"
        
        # Add header
        self.add_company_header(workbook, branding, "AP Register")
        
        # Add data table
        start_row = 5
        headers = ["Invoice No.", "Date", "Vendor", "Amount", "Tax", "Outstanding", "Status"]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            sheet.cell(row=start_row, column=col, value=header)
        
        # Write data
        current_row = start_row + 1
        total_amount = 0
        total_outstanding = 0
        
        # Handle both list of invoices and dict with invoices
        if isinstance(data, list):
            invoices = data
        elif isinstance(data, dict) and 'invoices' in data:
            invoices = data['invoices']
        else:
            invoices = []
        
        for invoice in invoices:
            invoice_no = invoice.get('invoice_no', '')
            date = self.format_date(invoice.get('invoice_date'))
            vendor = invoice.get('vendor_name', 'Unknown')
            amount = float(invoice.get('net_amt', invoice.get('invoice_amt', 0)))
            tax = float(invoice.get('tax_amt', 0))
            outstanding = float(invoice.get('outstanding', 0))
            status = invoice.get('status', 'Unknown')
            
            # Write row
            sheet.cell(row=current_row, column=1, value=invoice_no)
            sheet.cell(row=current_row, column=2, value=date)
            sheet.cell(row=current_row, column=3, value=vendor)
            sheet.cell(row=current_row, column=4, value=amount)
            sheet.cell(row=current_row, column=5, value=tax)
            sheet.cell(row=current_row, column=6, value=outstanding)
            sheet.cell(row=current_row, column=7, value=status)
            
            # Update totals
            total_amount += amount
            total_outstanding += outstanding
            
            current_row += 1
        
        # Add totals row
        sheet.cell(row=current_row, column=1, value="TOTAL").font = Font(bold=True)
        sheet.cell(row=current_row, column=4, value=total_amount).font = Font(bold=True)
        sheet.cell(row=current_row, column=6, value=total_outstanding).font = Font(bold=True)
        
        # Apply branding
        self.apply_branding(workbook, branding)
        
        # Add footer
        self.add_footer(workbook, branding)
        
        return workbook
    
    def render_dso_template(self, data: Any, branding: Dict[str, Any]) -> openpyxl.Workbook:
        """Render DSO report template"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "DSO Report"
        
        # Add header
        self.add_company_header(workbook, branding, "Days Sales Outstanding (DSO) Report")
        
        # Add summary data
        start_row = 5
        
        # Handle both list of invoices and dict with invoices
        if isinstance(data, list):
            # For DSO, we need to calculate from raw invoice data
            # This is a simplified version - in a real implementation, 
            # you'd calculate DSO from the invoice data
            dso_data = {
                'current_dso': 0,
                'previous_dso': 0,
                'dso_change': 0,
                'breakdown': {}
            }
        elif isinstance(data, dict) and 'dso_data' in data:
            dso_data = data['dso_data']
        else:
            dso_data = {}
        
        # DSO calculation summary
        current_dso = dso_data.get('current_dso', 0)
        previous_dso = dso_data.get('previous_dso', 0)
        dso_change = dso_data.get('dso_change', 0)
        
        sheet.cell(row=start_row, column=1, value="Current DSO").font = Font(bold=True)
        sheet.cell(row=start_row, column=2, value=current_dso)
        
        sheet.cell(row=start_row + 1, column=1, value="Previous DSO").font = Font(bold=True)
        sheet.cell(row=start_row + 1, column=2, value=previous_dso)
        
        sheet.cell(row=start_row + 2, column=1, value="DSO Change").font = Font(bold=True)
        sheet.cell(row=start_row + 2, column=2, value=dso_change)
        
        # Add detailed breakdown if available
        if 'breakdown' in dso_data:
            breakdown = dso_data['breakdown']
            sheet.cell(row=start_row + 4, column=1, value="DSO Breakdown").font = Font(bold=True, size=12)
            
            breakdown_headers = ["Period", "AR Balance", "Revenue", "DSO"]
            for col, header in enumerate(breakdown_headers, 1):
                sheet.cell(row=start_row + 5, column=col, value=header)
            
            current_row = start_row + 6
            for period, values in breakdown.items():
                sheet.cell(row=current_row, column=1, value=period)
                sheet.cell(row=current_row, column=2, value=values.get('ar_balance', 0))
                sheet.cell(row=current_row, column=3, value=values.get('revenue', 0))
                sheet.cell(row=current_row, column=4, value=values.get('dso', 0))
                current_row += 1
        
        # Apply branding
        self.apply_branding(workbook, branding)
        
        # Add footer
        self.add_footer(workbook, branding)
        
        return workbook
    
    def run(self, input_data: Any, params: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Generate Excel report
        
        Args:
            input_data: Raw invoice data or report data
            params: Configuration parameters
            
        Returns:
            File path and metadata
        """
        params = params or {}
        company_id = params.get('company_id')
        report_type = params.get('report_type', 'ap_aging')
        include_charts = params.get('include_charts', True)
        
        if not company_id:
            raise ValueError("Company ID is required for branding")
        
        # Load branding from database
        branding = self.load_branding(company_id)
        
        # Check if input_data is raw invoice data (list of dicts) or already formatted report data
        if isinstance(input_data, list) and len(input_data) > 0 and isinstance(input_data[0], dict):
            # Raw invoice data - convert to report format
            report_data = self._convert_invoices_to_report_data(input_data, report_type)
        else:
            # Already formatted report data
            report_data = input_data
        
        # Validate data using custom validation method
        if not self.validate_data(report_data):
            raise ValueError("Invalid data provided")
        
        # Select template based on report type
        if report_type == 'ap_aging':
            workbook = self.render_ap_aging_template(report_data, branding)
        elif report_type == 'ar_aging':
            workbook = self.render_ar_aging_template(report_data, branding)
        elif report_type == 'ap_register':
            workbook = self.render_ap_register_template(report_data, branding)
        elif report_type == 'dso':
            workbook = self.render_dso_template(report_data, branding)
        else:
            # Default to AP Aging
            workbook = self.render_ap_aging_template(report_data, branding)
        
        # Save file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{branding['company_name'].replace(' ', '_')}_{report_type.upper()}_{timestamp}.xlsx"
        file_path = f"./data/reports/{filename}"
        
        workbook.save(file_path)
        
        self.logger.info(f"Generated Excel report: {file_path}")
        
        return {
            'file_path': file_path,
            'branding_applied': True,
            'report_type': report_type,
            'branding': branding,
            'metadata': self.get_report_metadata(params)
        }
    
    def _convert_invoices_to_report_data(self, invoices: List[Dict], report_type: str) -> Dict:
        """
        Convert raw invoice data to report format
        
        Args:
            invoices: List of invoice dictionaries
            report_type: Type of report to generate
            
        Returns:
            Formatted report data
        """
        if report_type == 'ap_aging':
            return self._convert_to_ap_aging_format(invoices)
        elif report_type == 'ar_aging':
            return self._convert_to_ar_aging_format(invoices)
        elif report_type == 'ap_register':
            return self._convert_to_ap_register_format(invoices)
        else:
            return {'invoices': invoices}
    
    def _convert_to_ap_aging_format(self, invoices: List[Dict]) -> Dict:
        """Convert invoices to AP aging format"""
        from datetime import datetime, timedelta
        
        today = datetime.now().date()
        aging_data = []
        
        for invoice in invoices:
            # Calculate aging buckets
            due_date = self.parse_date(invoice.get('due_date'))
            if due_date:
                days_overdue = (today - due_date).days
            else:
                days_overdue = 0
            
            total_due = float(invoice.get('outstanding_amount', invoice.get('total_amount', 0)))
            
            # Assign to aging buckets
            if days_overdue <= 30:
                current_0_30 = total_due
                days_31_60 = 0
                days_61_90 = 0
                days_90_plus = 0
            elif days_overdue <= 60:
                current_0_30 = 0
                days_31_60 = total_due
                days_61_90 = 0
                days_90_plus = 0
            elif days_overdue <= 90:
                current_0_30 = 0
                days_31_60 = 0
                days_61_90 = total_due
                days_90_plus = 0
            else:
                current_0_30 = 0
                days_31_60 = 0
                days_61_90 = 0
                days_90_plus = total_due
            
            aging_data.append({
                'vendor_name': invoice.get('vendor_name', 'Unknown'),
                'invoice_no': invoice.get('invoice_number', ''),
                'due_date': invoice.get('due_date', ''),
                'total_due': total_due,
                'current_0_30': current_0_30,
                'days_31_60': days_31_60,
                'days_61_90': days_61_90,
                'days_90_plus': days_90_plus
            })
        
        return {'invoices': aging_data}
    
    def _convert_to_ar_aging_format(self, invoices: List[Dict]) -> Dict:
        """Convert invoices to AR aging format"""
        from datetime import datetime, timedelta
        
        today = datetime.now().date()
        aging_data = []
        
        for invoice in invoices:
            # Calculate aging buckets
            due_date = self.parse_date(invoice.get('due_date'))
            if due_date:
                days_overdue = (today - due_date).days
            else:
                days_overdue = 0
            
            total_due = float(invoice.get('outstanding_amount', invoice.get('total_amount', 0)))
            
            # Assign to aging buckets
            if days_overdue <= 30:
                current_0_30 = total_due
                days_31_60 = 0
                days_61_90 = 0
                days_90_plus = 0
            elif days_overdue <= 60:
                current_0_30 = 0
                days_31_60 = total_due
                days_61_90 = 0
                days_90_plus = 0
            elif days_overdue <= 90:
                current_0_30 = 0
                days_31_60 = 0
                days_61_90 = total_due
                days_90_plus = 0
            else:
                current_0_30 = 0
                days_31_60 = 0
                days_61_90 = 0
                days_90_plus = total_due
            
            aging_data.append({
                'customer_name': invoice.get('vendor_name', 'Unknown'),  # For AR, vendor becomes customer
                'invoice_no': invoice.get('invoice_number', ''),
                'due_date': invoice.get('due_date', ''),
                'total_due': total_due,
                'current_0_30': current_0_30,
                'days_31_60': days_31_60,
                'days_61_90': days_61_90,
                'days_90_plus': days_90_plus
            })
        
        return {'invoices': aging_data}
    
    def _convert_to_ap_register_format(self, invoices: List[Dict]) -> Dict:
        """Convert invoices to AP register format"""
        register_data = []
        
        for invoice in invoices:
            register_data.append({
                'invoice_no': invoice.get('invoice_number', ''),
                'invoice_date': invoice.get('invoice_date', ''),
                'vendor_name': invoice.get('vendor_name', 'Unknown'),
                'invoice_amt': float(invoice.get('total_amount', 0)),
                'tax_amt': float(invoice.get('tax_amount', 0)),
                'net_amt': float(invoice.get('subtotal_amount', 0)),
                'outstanding': float(invoice.get('outstanding_amount', 0)),
                'status': invoice.get('payment_status', 'Unknown')
            })
        
        return {'invoices': register_data}
    
    def validate_data(self, data: Any) -> bool:
        """
        Validate data before rendering - override base validation to handle raw invoice data
        """
        if not data:
            self.logger.error("No data provided for template")
            return False
        
        # For raw invoice data (list of dicts), basic validation is sufficient
        if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
            return True
        
        # For formatted report data, check for required structure
        if isinstance(data, dict) and 'invoices' in data:
            return True
        
        self.logger.error("Invalid data format - expected list of invoices or dict with invoices")
        return False
    
    def parse_date(self, date_str: str) -> Optional[datetime.date]:
        """Parse date string to date object"""
        if not date_str:
            return None
        
        try:
            from datetime import datetime
            # Try common date formats
            for fmt in ['%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y', '%m/%d/%Y']:
                try:
                    return datetime.strptime(date_str, fmt).date()
                except ValueError:
                    continue
            return None
        except Exception:
            return None
