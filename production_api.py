"""
PRODUCTION API - FINAL VERSION
Matches existing database schema exactly
"""

from fastapi import FastAPI, UploadFile, File, HTTPException, Depends, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel, EmailStr
from typing import Dict, Any, Optional, List
from pathlib import Path
from datetime import datetime, timedelta
import uuid
import os
import sys
import shutil
try:
    import jwt  # PyJWT
except ImportError:
    print("ERROR: PyJWT not installed. Run: pip install PyJWT")
    sys.exit(1)
import bcrypt

sys.path.insert(0, os.getcwd())

from sqlalchemy import create_engine, Column, String, DateTime, Integer, Boolean, Text, Numeric
from sqlalchemy.dialects.postgresql import JSONB, ARRAY
from sqlalchemy.orm import declarative_base, sessionmaker, Session
from sqlalchemy.pool import QueuePool

# Import existing components
from processing_layer.document_processing.document_processing_service import DocumentProcessingService
from processing_layer.document_processing.parsers.universal_docling_parser import UniversalDoclingParser
from processing_layer.document_processing.parsers.csv_parser import CSVParser
from intelligence_layer.routing.router_prompt_integrator import RouterPromptIntegrator
from intelligence_layer.parsing.enhanced_intent_parser import EnhancedIntentParser
from intelligence_layer.orchestration.workflow_planner_agent import WorkflowPlannerAgent
from processing_layer.agents.accounts_payable.ap_aging_agent import APAgingAgent
from processing_layer.agents.accounts_payable.ap_register_agent import APRegisterAgent
from processing_layer.agents.accounts_payable.ap_overdue_agent import APOverdueAgent
from processing_layer.agents.accounts_payable.ap_duplicate_agent import APDuplicateAgent
from processing_layer.agents.accounts_receivable.ar_aging_agent import ARAgingAgent
from processing_layer.agents.accounts_receivable.ar_register_agent import ARRegisterAgent
from processing_layer.agents.accounts_receivable.ar_collection_agent import ARCollectionAgent
from processing_layer.agents.accounts_receivable.dso_agent import DSOAgent
from processing_layer.workflows.nodes.base_node import NodeRegistry
from data_layer.database.database_manager import get_database
from shared.llm.groq_client import get_groq_client
from shared.config.logging_config import get_logger

logger = get_logger(__name__)

# ============================================================================
# DATABASE MODELS - EXACT MATCH TO YOUR SCHEMA
# ============================================================================

Base = declarative_base()

class User(Base):
    """Matches: public.users"""
    __tablename__ = "users"
    
    id = Column(String(36), primary_key=True)
    email = Column(String(255), unique=True, nullable=False)
    password_hash = Column(String(255), nullable=False)
    full_name = Column(String(255))
    company_id = Column(String(36), nullable=False)
    is_active = Column(Boolean)
    created_at = Column(DateTime)
    last_login = Column(DateTime)

class CompanySetupUpdate(BaseModel):
    primary_color: Optional[str] = None
    secondary_color: Optional[str] = None
    accent_color: Optional[str] = None
    default_currency: Optional[str] = None
    company_aliases: Optional[List[str]] = None

class Company(Base):
    """Matches: public.companies"""
    __tablename__ = "companies"
    
    id = Column(String(36), primary_key=True)
    name = Column(String(255), nullable=False)
    tax_id = Column(String(50))
    registration_number = Column(String(50))
    email = Column(String(255))
    phone = Column(String(50))
    website = Column(String(255))
    address_line1 = Column(String(255))
    address_line2 = Column(String(255))
    city = Column(String(100))
    state = Column(String(100))
    country = Column(String(100))
    postal_code = Column(String(20))
    logo_url = Column(String(500))
    company_aliases = Column(ARRAY(Text), nullable=True)
    primary_color = Column(String(7), default='#1976D2')
    secondary_color = Column(String(7), default='#424242')
    currency = Column(String(3), default='INR')
    fiscal_year_start = Column(String(5))
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime)
    updated_at = Column(DateTime)

class Document(Base):
    """Matches: public.documents"""
    __tablename__ = "documents"
    
    id = Column(String(36), primary_key=True)
    company_id = Column(String(36), nullable=False)
    file_name = Column(String(255), nullable=False)
    file_path = Column(String(500))
    file_type = Column(String(10))
    file_size = Column(Integer)
    document_type = Column(String(50))
    document_number = Column(String(100))
    document_date = Column(DateTime)
    category = Column(String(50))
    docling_parsed_data = Column(JSONB)
    canonical_data = Column(JSONB)
    status = Column(String(50), default='pending')
    confidence_score = Column(Numeric(3, 2))
    vendor_name = Column(String(255))
    customer_name = Column(String(255))
    grand_total = Column(Numeric(15, 2), default=0.0)
    tax_total = Column(Numeric(15, 2), default=0.0)
    paid_amount = Column(Numeric(15, 2), default=0.0)
    outstanding = Column(Numeric(15, 2), default=0.0)
    uploaded_at = Column(DateTime)
    parsed_at = Column(DateTime)
    processed_at = Column(DateTime)
    created_at = Column(DateTime)
    updated_at = Column(DateTime)

class Workflow(Base):
    """Matches: public.workflows"""
    __tablename__ = "workflows"
    
    id = Column(String(36), primary_key=True)
    name = Column(String(255), nullable=False)
    type = Column(String(50), nullable=False)
    status = Column(String(50), nullable=False, default='planned')
    company_id = Column(String(50), nullable=False)
    user_id = Column(String(50), nullable=False)
    query = Column(Text)
    domain = Column(String(50))
    report_type = Column(String(100))
    workflow_definition = Column(JSONB, nullable=False)
    execution_result = Column(JSONB)
    created_at = Column(DateTime, nullable=False)
    started_at = Column(DateTime)
    completed_at = Column(DateTime)
    output_file_path = Column(String(500))
    execution_time_ms = Column(Integer)
    error_message = Column(Text)

# ============================================================================
# CONFIGURATION
# ============================================================================

class Config:
    DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://postgres:postgres@localhost:5432/financial_automation")
    UPLOAD_DIR = Path("./data/uploads")
    OUTPUT_DIR = Path("./data/reports")
    API_HOST = "0.0.0.0"
    API_PORT = 8000
    JWT_SECRET = os.getenv("JWT_SECRET", "your-secret-key-change-in-production")
    TOKEN_EXPIRE_HOURS = 24

config = Config()
config.UPLOAD_DIR.mkdir(exist_ok=True, parents=True)
config.OUTPUT_DIR.mkdir(exist_ok=True, parents=True)

# ============================================================================
# DATABASE
# ============================================================================

engine = create_engine(config.DATABASE_URL, poolclass=QueuePool, pool_size=10)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ============================================================================
# INITIALIZE SERVICES
# ============================================================================

db_manager = get_database()
docling_parser = UniversalDoclingParser()
csv_parser = CSVParser()
llm_client = get_groq_client("accurate")
router = RouterPromptIntegrator()
intent_parser = EnhancedIntentParser(llm_client=llm_client)
workflow_planner = WorkflowPlannerAgent(llm_client=llm_client)

AGENTS = {
    'ap_aging': APAgingAgent(),
    'ap_register': APRegisterAgent(),
    'ap_overdue': APOverdueAgent(),
    'ap_duplicate': APDuplicateAgent(),
    'ar_aging': ARAgingAgent(),
    'ar_register': ARRegisterAgent(),
    'ar_collection': ARCollectionAgent(),
    'dso': DSOAgent()
}

logger.info(f" System initialized with {len(AGENTS)} agents")

# ============================================================================
# SECURITY
# ============================================================================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str) -> str:
    payload = {"sub": user_id, "exp": datetime.utcnow() + timedelta(hours=config.TOKEN_EXPIRE_HOURS)}
    return jwt.encode(payload, config.JWT_SECRET, algorithm="HS256")

def decode_token(token: str) -> dict:
    try:
        return jwt.decode(token, config.JWT_SECRET, algorithms=["HS256"])
    except:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_current_user(authorization: str = Header(...), db: Session = Depends(get_db)) -> User:
    token = authorization.replace("Bearer ", "")
    payload = decode_token(token)
    user = db.query(User).filter(User.id == payload['sub']).first()
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user

# ============================================================================
# MODELS
# ============================================================================

class UserRegister(BaseModel):
    email: EmailStr
    password: str
    full_name: str
    company_name: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class ChatQuery(BaseModel):
    query: str

# ============================================================================
# FASTAPI APP
# ============================================================================

app = FastAPI(title="Financial Automation System", version="3.1.0", docs_url="/docs")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"]
)

# Mount static files directory
from fastapi.staticfiles import StaticFiles
app.mount("/static", StaticFiles(directory="./data"), name="static")

# ============================================================================
# ENDPOINTS
# ============================================================================

@app.get("/")
async def root():
    return {
        "service": "Financial Automation System",
        "version": "3.1.0",
        "status": "operational",
        "features": {
            "agents": len(AGENTS),
            "nodes": len(NodeRegistry.get_all_nodes())
        }
    }

@app.get("/api/v1/debug/documents")
async def debug_documents(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Debug endpoint to check documents"""
    
    from sqlalchemy import text
    
    # Check what's in database
    query = text("""
        SELECT 
            id, file_name, category, vendor_name, customer_name, 
            grand_total, status
        FROM documents
        WHERE company_id = :company_id
        ORDER BY created_at DESC
        LIMIT 10
    """)
    
    result = db.execute(query, {"company_id": current_user.company_id})
    docs = [dict(row._mapping) for row in result]
    
    # Check categories
    category_query = text("""
        SELECT category, COUNT(*) as count
        FROM documents
        WHERE company_id = :company_id
        GROUP BY category
    """)
    
    cat_result = db.execute(category_query, {"company_id": current_user.company_id})
    categories = [dict(row._mapping) for row in cat_result]
    
    return {
        "company_id": current_user.company_id,
        "documents": docs,
        "categories": categories
    }

@app.get("/health")
async def health():
    return {
        "status": "healthy",
        "timestamp": datetime.utcnow().isoformat(),
        "agents": len(AGENTS),
        "nodes": len(NodeRegistry.get_all_nodes())
    }

@app.post("/api/v1/auth/register")
async def register(user_data: UserRegister, db: Session = Depends(get_db)):
    """Register new user"""
    
    if db.query(User).filter(User.email == user_data.email).first():
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = str(uuid.uuid4())
    company_id = str(uuid.uuid4())
    
    # Create company
    company = Company(
        id=company_id,
        name=user_data.company_name,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow()
    )
    
    # Create user
    user = User(
        id=user_id,
        company_id=company_id,
        email=user_data.email,
        password_hash=hash_password(user_data.password),
        full_name=user_data.full_name,
        is_active=True,
        created_at=datetime.utcnow()
    )
    
    db.add(company)
    db.add(user)
    db.commit()
    
    token = create_token(user_id)
    
    logger.info(f" User registered: {user_data.email} - {user_data.company_name}")
    
    return {
        "status": "success",
        "access_token": token,
        "token_type": "bearer",
        "user_id": user_id,
        "company_id": company_id,
        "company_name": user_data.company_name
    }

@app.post("/api/v1/auth/login")
async def login(credentials: UserLogin, db: Session = Depends(get_db)):
    """Login"""
    
    user = db.query(User).filter(User.email == credentials.email).first()
    
    if not user or not verify_password(credentials.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_token(user.id)
    company = db.query(Company).filter(Company.id == user.company_id).first()
    
    return {
        "status": "success",
        "access_token": token,
        "token_type": "bearer",
        "user_id": user.id,
        "company_id": user.company_id,
        "company_name": company.name if company else "Unknown"
    }

@app.put("/api/v1/company/setup")
async def update_company_setup(
    setup_data: CompanySetupUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update company branding and settings"""
    
    company = db.query(Company).filter(Company.id == current_user.company_id).first()
    
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    # Update fields if provided
    if 'primary_color' in setup_data:
        company.primary_color = setup_data['primary_color']
    if 'secondary_color' in setup_data:
        company.secondary_color = setup_data['secondary_color']
    if 'accent_color' in setup_data:
        company.accent_color = setup_data.get('accent_color')
    if 'default_currency' in setup_data:
        company.currency = setup_data['default_currency']
    if setup_data.company_aliases is not None:
        #  sanitize input
        company.company_aliases = [
            name.strip()
            for name in setup_data.company_aliases
            if name.strip()
        ]
    
    company.updated_at = datetime.utcnow()
    db.commit()
    
    logger.info(f" Company setup updated: {company.name}")
    
    return {
        "status": "success",
        "message": "Company setup updated",
        "company": {
            "name": company.name,
            "aliases": company.company_aliases,
            "logo_url": company.logo_url,
            "primary_color": company.primary_color,
            "secondary_color": company.secondary_color,
            "currency": company.currency
        }
    }

@app.post("/api/v1/company/logo")
async def upload_company_logo(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Upload company logo (PNG, JPG, JPEG, SVG)"""
    
    try:
        # Validate file type
        allowed_extensions = {'.png', '.jpg', '.jpeg', '.svg', '.webp'}
        file_ext = Path(file.filename).suffix.lower()
        
        if file_ext not in allowed_extensions:
            raise HTTPException(
                status_code=400, 
                detail=f"Invalid file type. Allowed: PNG, JPG, JPEG, SVG, WEBP"
            )
        
        # Validate file size (max 5MB)
        file.file.seek(0, 2)  # Seek to end
        file_size = file.file.tell()  # Get position (size)
        file.file.seek(0)  # Reset to start
        
        max_size = 5 * 1024 * 1024  # 5MB
        if file_size > max_size:
            raise HTTPException(
                status_code=400,
                detail=f"File too large. Max size: 5MB. Your file: {file_size / 1024 / 1024:.2f}MB"
            )
        
        # Create logos directory
        logo_dir = Path("./data/logos")
        logo_dir.mkdir(exist_ok=True, parents=True)
        
        # Save logo with company ID prefix
        logo_filename = f"{current_user.company_id}_logo{file_ext}"
        logo_path = logo_dir / logo_filename
        
        # Save file
        with logo_path.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        # Update company
        company = db.query(Company).filter(Company.id == current_user.company_id).first()
        company.logo_url = f"/static/logos/{logo_filename}"
        company.updated_at = datetime.utcnow()
        db.commit()
        
        logger.info(f" Logo uploaded for: {company.name} ({file_size / 1024:.2f} KB)")
        
        return {
            "status": "success",
            "logo_url": company.logo_url,
            "file_size_kb": round(file_size / 1024, 2),
            "file_type": file_ext,
            "message": "Logo uploaded successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Logo upload failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/v1/documents/upload")
async def upload_document(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Upload invoice"""
    
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{timestamp}_{file.filename}"
        filepath = config.UPLOAD_DIR / filename
        
        with filepath.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        company = db.query(Company).filter(Company.id == current_user.company_id).first()
        
        file_ext = Path(file.filename).suffix.lower()
        parser = csv_parser if file_ext == '.csv' else docling_parser
        
        # Get DatabaseManager instance (not just connection)
        from data_layer.database.database_manager import get_database
        db_manager_instance = get_database()
        
        processor = DocumentProcessingService(
            db_session=db_manager_instance,
            docling_parser=parser,
            company_id=current_user.company_id,
            user_company_name=company.name  # For intelligent classification
        )
        
        result = processor.process_upload(
            file_path=str(filepath),
            file_name=file.filename
        )
        
        if not result.get("success"):
            raise HTTPException(status_code=500, detail="Processing failed")
        
        logger.info(f" Document processed: {result['document_id']}")
        
        return {
            "status": "success",
            "document_id": result['document_id'],
            "category": result.get('invoice_category'),
            "party": result.get('party_info'),
            "extracted_data": result.get('extracted_data')
        }
        
    except Exception as e:
        logger.error(f"Upload failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/v1/documents")
async def list_documents(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    limit: int = 100
):
    """List uploaded documents"""
    
    documents = db.query(Document).filter(
        Document.company_id == current_user.company_id
    ).order_by(Document.uploaded_at.desc()).limit(limit).all()
    
    return {
        "status": "success",
        "count": len(documents),
        "documents": [
            {
                "id": doc.id,
                "file_name": doc.file_name,
                "document_type": doc.document_type,
                "category": doc.category,
                "uploaded_at": doc.uploaded_at.isoformat() if doc.uploaded_at else None
            }
            for doc in documents
        ]
    }

@app.post("/api/v1/chat/query")
async def chat_query(
    query_data: ChatQuery,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Process natural language query and show workflow"""
    
    query = query_data.query
    start_time = datetime.now()
    
    try:
        logger.info(f"[CHAT] Query: {query}")
        
        company = db.query(Company).filter(Company.id == current_user.company_id).first()
        
        # Route query
        routing_result = router.process_query(query)
        domain = routing_result.get('domain', 'APLayer')
        
        # Parse intent
        intent_result = intent_parser.parse(query)
        report_type = intent_result.get('report_type')
        variables = intent_result.get('variables', {})
        
        # Get agent
        agent = AGENTS.get(report_type)
        if not agent:
            raise HTTPException(status_code=400, detail=f"Unknown report type: {report_type}")
        
        agent_name = agent.__class__.__name__
        logger.info(f"   Agent: {agent_name}")
        
        # Build workflow - THIS SHOWS THE VISUAL PIPELINE
        logger.info("[WORKFLOW] Building visual workflow...")
        workflow_def = workflow_planner.execute(
            input_data=query,
            params={'report_type': report_type, 'domain': domain, **variables}
        )
        
        workflow_steps = workflow_def.get('workflow', {}).get('steps', [])
        workflow_edges = workflow_def.get('workflow', {}).get('edges', [])
        
        # Create workflow record BEFORE execution
        workflow_id = str(uuid.uuid4())
        
        workflow = Workflow(
            id=workflow_id,
            name=f"{report_type.upper()} Report",
            type=report_type,
            status='executing',
            company_id=current_user.company_id,
            user_id=current_user.id,
            query=query,
            domain=domain,
            report_type=report_type,
            workflow_definition={
                'nodes': workflow_steps,
                'edges': workflow_edges
            },
            created_at=datetime.utcnow(),
            started_at=start_time
        )
        
        db.add(workflow)
        db.commit()
        
        # Format nodes for visual display
        visual_nodes = []
        for idx, step in enumerate(workflow_steps):
            visual_nodes.append({
                "id": step.get('id', f"node_{idx}"),
                "type": step.get('type'),
                "name": step.get('name', step.get('type')),
                "position": {"x": 100 + (idx * 250), "y": 100},
                "status": "pending",
                "data": {
                    "label": step.get('name', step.get('type')),
                    "params": step.get('params', {}),
                    "description": step.get('description', '')
                }
            })
        
        logger.info(f"[WORKFLOW] Created workflow with {len(visual_nodes)} nodes")
        logger.info(f"[WORKFLOW] Workflow ID: {workflow_id}")
        
        # NOW EXECUTE - this happens in background but user already sees workflow
        try:
            params = {
                'user_id': current_user.id,
                'company_id': current_user.company_id,
                'company_name': company.name,
                **variables
            }
            
            result = agent.execute(input_data=None, params=params)
            
            if result.get('status') != 'success':
                workflow.status = 'failed'
                workflow.error_message = result.get('message', 'Execution failed')
                db.commit()
            else:
                file_path = result.get('file_path') or result.get('data', {}).get('file_path')
                
                workflow.status = 'completed'
                workflow.completed_at = datetime.utcnow()
                workflow.execution_time_ms = int((datetime.now() - start_time).total_seconds() * 1000)
                workflow.output_file_path = file_path
                workflow.execution_result = result
                db.commit()
                
                # Update node statuses to completed
                for node in visual_nodes:
                    node['status'] = 'completed'
        
        except Exception as exec_error:
            logger.error(f"Execution error: {exec_error}")
            workflow.status = 'failed'
            workflow.error_message = str(exec_error)
            db.commit()
            
            # Mark nodes as failed
            for node in visual_nodes:
                node['status'] = 'failed'
        
        execution_time = int((datetime.now() - start_time).total_seconds() * 1000)
        
        logger.info(f" Complete - {execution_time}ms")
        
        return {
            "status": "success",
            "workflow": {
                "id": workflow_id,
                "name": workflow.name,
                "query": query,
                "agent": agent_name,
                "domain": domain,
                "report_type": report_type,
                "status": workflow.status,
                "nodes": visual_nodes,
                "edges": workflow_edges,
                "created_at": workflow.created_at.isoformat()
            },
            "report": {
                "file_path": workflow.output_file_path,
                "download_url": f"/api/v1/reports/download/{Path(workflow.output_file_path).name}" if workflow.output_file_path else None
            },
            "execution": {
                "time_ms": execution_time,
                "nodes_executed": len(visual_nodes)
            }
        }
        
    except Exception as e:
        logger.error(f"Query failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/v1/workflows")
async def list_workflows(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    limit: int = 50
):
    """List workflows"""
    
    workflows = db.query(Workflow).filter(
        Workflow.company_id == current_user.company_id
    ).order_by(Workflow.created_at.desc()).limit(limit).all()
    
    return {
        "status": "success",
        "count": len(workflows),
        "workflows": [
            {
                "id": wf.id,
                "name": wf.name,
                "query": wf.query,
                "type": wf.type,
                "status": wf.status,
                "created_at": wf.created_at.isoformat(),
                "output_file": wf.output_file_path
            }
            for wf in workflows
        ]
    }

@app.get("/api/v1/reports/download/{filename}")
async def download_report(filename: str):
    """Download report"""
    file_path = config.OUTPUT_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(file_path, filename=filename)

@app.get("/api/v1/agents")
async def list_agents():
    return {
        "status": "success",
        "count": len(AGENTS),
        "agents": {key: agent.__class__.__name__ for key, agent in AGENTS.items()}
    }

@app.get("/api/v1/nodes")
async def list_nodes():
    nodes = NodeRegistry.get_all_nodes()
    return {
        "status": "success",
        "count": len(nodes),
        "nodes": nodes
    }

@app.post("/api/v1/reports/ap-register/simple")
async def generate_simple_ap_register(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate simple AP Register from documents table"""
    
    try:
        from sqlalchemy import text
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Query documents
        query = text("""
            SELECT 
                id,
                file_name,
                document_number,
                document_date,
                vendor_name,
                customer_name,
                grand_total,
                tax_total,
                outstanding,
                category,
                status
            FROM documents
            WHERE company_id = :company_id
            AND category = 'purchase'
            ORDER BY document_date DESC NULLS LAST, file_name
        """)
        
        result = db.execute(query, {"company_id": current_user.company_id})
        invoices = result.fetchall()
        
        if not invoices:
            return {
                "status": "success",
                "message": "No purchase invoices found",
                "record_count": 0
            }
        
        # Create Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AP Register"
        
        # Header
        headers = ["Invoice Number", "Date", "Vendor", "Amount", "Tax", "Outstanding", "Status"]
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        total_amount = 0
        for inv in invoices:
            ws.append([
                inv.document_number or inv.file_name,
                inv.document_date.strftime("%Y-%m-%d") if inv.document_date else "",
                inv.vendor_name or "Unknown",
                float(inv.grand_total or 0),
                float(inv.tax_total or 0),
                float(inv.outstanding or inv.grand_total or 0),
                inv.status or "pending"
            ])
            total_amount += float(inv.grand_total or 0)
        
        # Total row
        ws.append([])
        total_row = ws.max_row + 1
        ws.cell(total_row, 1, "TOTAL")
        ws.cell(total_row, 4, total_amount)
        
        for cell in ws[total_row]:
            cell.font = Font(bold=True)
        
        # Auto-width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2
        
        # Save
        filename = f"ap_register_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = config.OUTPUT_DIR / filename
        
        wb.save(filepath)
        
        logger.info(f" Generated AP Register: {filename}")
        
        return {
            "status": "success",
            "file_path": str(filepath),
            "download_url": f"/api/v1/reports/download/{filename}",
            "record_count": len(invoices),
            "total_amount": total_amount
        }
        
    except Exception as e:
        logger.error(f"Report generation failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================================
# RUN
# ============================================================================

if __name__ == "__main__":
    import uvicorn
    
    print("\n" + "=" * 80)
    print("🚀 FINANCIAL AUTOMATION SYSTEM - READY")
    print("=" * 80)
    print(f"\n🌐 Server: http://{config.API_HOST}:{config.API_PORT}")
    print(f"📖 Docs:   http://{config.API_HOST}:{config.API_PORT}/docs\n")
    print("=" * 80 + "\n")
    
    uvicorn.run(app, host=config.API_HOST, port=config.API_PORT, reload=False)